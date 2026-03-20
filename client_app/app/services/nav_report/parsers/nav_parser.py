import io, re
import openpyxl

ALIASES = {
  "unit": ["单位净值", "基金单位净值", "最新单位净值", "今日单位净值", "份额净值", "净值"],
  "cum": ["累计净值", "累计单位净值", "复权净值"],
  "day": ["净值日增长率(%)", "净值日增长率", "日增长率(%)", "日涨幅(%)", "净值日涨幅(%)"],
  "week": ["净值周增长率(%)", "净值周增长率", "周增长率(%)", "周涨幅(%)", "净值周涨幅(%)"],
  "month": ["净值月增长率(%)", "净值月增长率", "月增长率(%)", "月涨幅(%)", "净值月涨幅(%)"],
  "ytd": ["今年来收益(%)", "年初至今(%)", "年初至今收益(%)", "净值年增长率(%)"]
}

def _to_float(x):
    if x is None: return None
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip().replace(",", "")
    if s == "": return None
    if s.endswith("%"):
        try: return float(s[:-1]) / 100.0
        except: return None
    try: return float(s)
    except: return None



def _scan_sheet_like(max_row, max_col, cell_get):
    hits = {k: [] for k in ALIASES.keys()}

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = cell_get(r, c)
            if v is None:
                continue
            txt = str(v).strip()
            for key, names in ALIASES.items():
                if any(txt == n for n in names):
                    hits[key].append((r, c, txt))

    def pick_right_value(hit_list):
        for (r, c, _) in hit_list:
            for dc in (1, 2, 3):
                vv = cell_get(r, c + dc)
                fv = _to_float(vv)
                if fv is not None:
                    return fv
        return None

    def fuzzy_find(keys):
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = cell_get(r, c)
                if v is None:
                    continue
                t = str(v).strip()
                if any(k in t for k in keys):
                    for dc in (1, 2, 3):
                        fv = _to_float(cell_get(r, c + dc))
                        if fv is not None:
                            return fv
        return None

    unit_nav = pick_right_value(hits["unit"])
    cum_nav = pick_right_value(hits["cum"])
    day = pick_right_value(hits["day"])
    week = pick_right_value(hits["week"])
    month = pick_right_value(hits["month"])
    ytd = pick_right_value(hits["ytd"])

    if unit_nav is None: unit_nav = fuzzy_find(["单位净值", "基金单位净值", "最新单位净值", "今日单位净值", "份额净值", "净值"])
    if cum_nav is None: cum_nav = fuzzy_find(["累计净值", "累计单位净值", "复权净值"])
    if day is None: day = fuzzy_find(["日增长率","日涨幅"])
    if week is None: week = fuzzy_find(["周增长率","周涨幅"])
    if month is None: month = fuzzy_find(["月增长率","月涨幅"])
    if ytd is None: ytd = fuzzy_find(["年初至今","今年来","年增长率"])

    nav = unit_nav
    return {
      "nav": nav,
      "unit_nav": unit_nav,
      "cum_nav": cum_nav,
      "day_pct": day,
      "week_pct": week,
      "month_pct": month,
      "file_ytd_pct": ytd,
      "year_pct": ytd
    }

def parse_metrics(xlsx_bytes: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        sh = wb[wb.sheetnames[0]]
        return _scan_sheet_like(sh.max_row, sh.max_column, lambda r, c: sh.cell(r, c).value)
    except Exception:
        pass

    try:
        import xlrd
        book = xlrd.open_workbook(file_contents=xlsx_bytes)
        sh = book.sheet_by_index(0)
        return _scan_sheet_like(sh.nrows, sh.ncols, lambda r, c: sh.cell_value(r - 1, c - 1) if 1 <= r <= sh.nrows and 1 <= c <= sh.ncols else None)
    except Exception:
        raise


import datetime as _dt

def _pick_date_from_filename(filename: str):
    m = re.search(r'(\d{4})[-_/\.](\d{2})[-_/\.](\d{2})', filename)
    if not m:
        return None
    return f"{m.group(1)}{m.group(2)}{m.group(3)}"

def _norm_product_key(name: str):
    name = (name or "").strip()
    # trim ugly tail like __xxxx__uid...__filename
    name = re.sub(r'__[^_]{6,}__uid[^_]+__.*$', '', name)
    name = re.sub(r'\s+', '', name)
    return name

def _extract_nav_payload(wb, filename: str, received_ymd: str):
    """
    Return dict: {product_key, valuation_date, nav}
    valuation_date priority:
      1) any explicit ymd found in workbook text cells (first hit)
      2) filename date (YYYYMMDD)
      3) fallback to received_ymd (already pre-processed as 'workday')
    """
    # 1) try read any date-like cell
    val_date = None
    try:
        for ws in wb.worksheets[:3]:
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for v in row:
                    if not v:
                        continue
                    t = str(v)
                    # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD
                    m = re.search(r'(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})', t)
                    if m:
                        y,mo,da = m.group(1), int(m.group(2)), int(m.group(3))
                        val_date = f"{y}{mo:02d}{da:02d}"
                        raise StopIteration
                    # MM.DD (assume same year as filename if has, else received year)
                    m2 = re.search(r'(^|[^\d])(\d{1,2})\.(\d{1,2})([^\d]|$)', t)
                    if m2 and not val_date:
                        mo,da = int(m2.group(2)), int(m2.group(3))
                        base = _pick_date_from_filename(filename) or received_ymd
                        y = int(base[:4])
                        val_date = f"{y}{mo:02d}{da:02d}"
                        raise StopIteration
    except StopIteration:
        pass

    if not val_date:
        val_date = _pick_date_from_filename(filename) or received_ymd

    # product_key from filename like "SXZ218_大可放心8号私募证券投资基金估值表.xlsx"
    product_key = filename
    product_key = re.sub(r'\.xlsx$', '', product_key, flags=re.I)
    product_key = re.sub(r'^\d{4}[-_/\.]\d{2}[-_/\.]\d{2}_', '', product_key)  # drop leading date_
    product_key = re.sub(r'_?估值表.*$', '', product_key)  # drop suffix
    product_key = product_key.replace("__", "_")
    product_key = _norm_product_key(product_key)

    # nav: find label "基金单位净值" nearby
    nav = None
    for ws in wb.worksheets:
        for r in range(1, 80):
            row = [ws.cell(r, c).value for c in range(1, 15)]
            row_s = [("" if v is None else str(v)).strip() for v in row]
            for idx, cell in enumerate(row_s):
                if "基金单位净值" in cell:
                    # look right cells
                    for j in range(idx+1, min(idx+6, len(row_s))):
                        x = row[j]
                        if x is None:
                            continue
                        try:
                            nav = float(str(x).replace(",", "").strip())
                            break
                        except:
                            continue
                    if nav is not None:
                        break
            if nav is not None:
                break
        if nav is not None:
            break

    return {"product_key": product_key, "valuation_date": val_date, "nav": nav}

