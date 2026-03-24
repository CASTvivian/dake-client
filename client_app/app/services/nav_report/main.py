
def _fdebug(msg: str):
    if FETCH_DEBUG:
        try:
            print("[FETCH_DEBUG]", msg, flush=True)
        except Exception:
            pass


def _resolve_rates(product_key: str, date_str: str, nav_today: float | None,
                   rate_day: float | None, rate_week: float | None, rate_month: float | None, rate_year: float | None):
    # Prefer provided rates; fallback to history-based calculation
    if rate_day is None:
        rate_day = pct_change(nav_today, get_prev_nav(product_key, date_str))
    if rate_week is None:
        rate_week = pct_change(nav_today, get_last_week_end_nav(product_key, date_str))
    if rate_month is None:
        rate_month = pct_change(nav_today, get_last_month_end_nav(product_key, date_str))
    if rate_year is None:
        rate_year = pct_change(nav_today, get_year_start_nav_A(product_key, date_str))
    return rate_day, rate_week, rate_month, rate_year

import csv
import datetime
from client_app.app.services.nav_report.utils.trading_calendar import build_calendar_from_env
import hashlib
from email.header import decode_header
import os
import logging

SUBJECT_HINT_KEYWORDS = [x.strip() for x in os.environ.get("SUBJECT_HINT_KEYWORDS","").split(",") if x.strip()]
NAV_FIELD_KEYWORDS = [
    "基金单位净值",
    "单位净值",
    "今日单位净值",
    "最新单位净值",
]

def _subject_is_candidate(subj: str) -> bool:
    if not subj:
        return False
    subj = str(subj)
    return any(k in subj for k in SUBJECT_HINT_KEYWORDS)

def _file_has_nav_field(path: str) -> bool:
    """
    内容识别：以文件内容是否出现 NAV_FIELD_KEYWORDS 为准。
    xlsx: 用 openpyxl 扫描前若干行列
    csv/txt: 直接文本搜索
    其他：尽量做 bytes 搜索（可能命不中，但不直接误判）
    """
    import os
    import pathlib
    fp = pathlib.Path(path)
    ext = fp.suffix.lower()

    # quick bytes scan (for csv/txt or fallback)
    try:
        b = fp.read_bytes()
        for kw in NAV_FIELD_KEYWORDS:
            if kw.encode("utf-8") in b:
                return True
    except Exception:
        pass

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            # 扫描最多前 3 个 sheet，每个 sheet 前 200 行、20 列
            for si, ws in enumerate(wb.worksheets[:3]):
                for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True), start=1):
                    for v in row:
                        if v is None:
                            continue
                        sv = str(v)
                        if any(kw in sv for kw in NAV_FIELD_KEYWORDS):
                            return True
            return False
        except Exception:
            return False

    # csv/txt 已做 bytes scan，走到这说明没命中
    return False

import re
import zipfile
import io

from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from client_app.app.services.nav_report.parsers.nav_parser import parse_metrics
from client_app.app.services.nav_report.utils.db import connect, init_db, upsert_nav, ensure_nav_series, get_prev_nav, get_last_week_end_nav, get_last_month_end_nav, get_year_start_nav_A, pct_change
from client_app.app.services.nav_report.utils.imap_fetch import fetch_attachments
from client_app.app.services.nav_report.utils.naming import product_key_from_filename, yyyymmdd_dir
from client_app.app.services.nav_report.utils import wecom_mail

load_dotenv()

PORT = int(os.getenv("PORT", "6002"))
DATA_DIR = os.getenv("DATA_DIR", "/data/services/nav_report_6002/data")

IMAP_HOST = os.getenv("IMAP_HOST", "")
IMAP_PORT = int(os.getenv("IMAP_PORT", "993"))
IMAP_USER = os.getenv("IMAP_USER", "")
IMAP_PASS = os.getenv("IMAP_PASS", "")
IMAP_FOLDER = os.getenv("IMAP_FOLDER", "INBOX")
MAIL_PROVIDER = os.getenv('MAIL_PROVIDER', 'imap').strip().lower()
MAIL_SINCE_HOURS = int(os.getenv("MAIL_SINCE_HOURS", "24"))
MAIL_UNSEEN_ONLY = os.getenv("MAIL_UNSEEN_ONLY", "true").lower() == "true"
MAIL_FROM_FILTER = os.getenv("MAIL_FROM_FILTER", "").strip() or None
MAIL_SUBJECT_KEYWORD = os.getenv("MAIL_SUBJECT_KEYWORD", "").strip() or None
MAIL_SUBJECT_REGEX = os.getenv("MAIL_SUBJECT_REGEX", "").strip() or None

YTD_BASE_MODE = os.getenv("YTD_BASE_MODE", "A").strip().upper()




def _to_ratio_pct_input(v):
    """Normalize percent-like inputs into ratio. 0.01 means 1%."""
    if v is None:
        return None
    try:
        if isinstance(v, str):
            t = v.strip().replace('%', '').replace(',', '')
            if t == '':
                return None
            v = float(t)
        v = float(v)
    except Exception:
        return None
    if abs(v) > 1:
        return v / 100.0
    return v

class RunReq(BaseModel):
    strict: bool = True
    target_val_date_only: bool = True
    folders: list[str] | None = None
    ingest_mode: str = 'append'  # append|replace
    lookback_days: int = 3
    recent_only: bool = False
    newer_than_hours: int = 6
    folder: str | None = None
    date_str: str = ""
    force: bool = False
    push: bool = True
    slot: str = ""


class ProductReq(BaseModel):
    code: str
    display_name: str | None = None


class ToggleReq(BaseModel):
    code: str
    enabled: bool


class SlotReq(BaseModel):
    slot: str = ""
    force: bool = True


def _pick_date(payload: Optional[RunReq], date_str: str) -> str:
    d = (payload.date_str if payload else date_str).strip()
    return d if len(d) == 8 else datetime.datetime.now().strftime("%Y%m%d")


def _pick_force(payload: Optional[RunReq], force: bool) -> bool:
    return payload.force if payload is not None else force

def _pick_lookback(payload: Optional[RunReq]) -> int:
    v = payload.lookback_days if payload is not None else _env_int("MAIL_LOOKBACK_DAYS", 3)
    try:
        v = int(v)
    except Exception:
        v = 3
    if v < 1:
        return 1
    if v > 14:
        return 14
    return v

def _pick_folder(payload: Optional[RunReq]) -> str:
    if payload is not None and payload.folder and str(payload.folder).strip():
        return str(payload.folder).strip()
    return IMAP_FOLDER


def _now_iso() -> str:
    return datetime.datetime.now().isoformat()


def _sha256(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _make_ingest_key(uid: str, filename: str, content: bytes) -> str:
    base = f"{uid}|{filename}|{_sha256(content)}".encode("utf-8")
    return _sha256(base)


def _decode_mime_filename(name: str) -> str:
    """Decode RFC2047 encoded-words in filename (=?UTF-8?...?=)."""
    try:
        parts = decode_header(name or "")
        out = []
        for b, enc in parts:
            if isinstance(b, bytes):
                out.append(b.decode(enc or "utf-8", errors="ignore"))
            else:
                out.append(str(b))
        return "".join(out).strip()
    except Exception:
        return (name or "").strip()

_ymd8_re = re.compile(r"(?<!\d)(20\d{6})(?!\d)")
_ymd_sep_re = re.compile(r"(20\d{2})[\-_/\.](\d{1,2})[\-_/\.](\d{1,2})")

def _is_valid_ymd(ymd: str) -> bool:
    try:
        datetime.datetime.strptime(ymd, "%Y%m%d")
        return True
    except Exception:
        return False


def _extract_ymd_from_filename(filename: str) -> str:
    fn = _decode_mime_filename(filename or "")
    m = _ymd8_re.search(fn)
    if m:
        ymd = m.group(1)
        return ymd if _is_valid_ymd(ymd) else ""
    m = _ymd_sep_re.search(fn)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        ymd = f"{int(y):04d}{int(mo):02d}{int(d):02d}"
        return ymd if _is_valid_ymd(ymd) else ""
    return ""


# CODEX_PRODUCT_ALLOWLIST_HELPERS
def _default_target_products() -> list[tuple[str, str]]:
    return [
        ("SBGZ87", "缙云"),
        ("STR134", "新力"),
        ("SXA927", "放10"),
        ("SXZ218", "放8"),
    ]


def _norm_code(code: str) -> str:
    return (code or "").strip().upper()


def _get_allow_codes() -> set[str]:
    raw = (os.getenv("PRODUCT_CODE_ALLOWLIST", "") or "").strip()
    return {_norm_code(x) for x in raw.split(",") if x.strip()}


def _get_allowlist_from_db(conn_obj) -> list[dict]:
    cur = conn_obj.cursor()
    cur.execute("SELECT code, display_name, enabled, updated_at FROM product_allowlist ORDER BY code")
    rows = cur.fetchall()
    return [
        {"code": r[0], "display_name": r[1] or "", "enabled": int(r[2]), "updated_at": r[3] or ""}
        for r in rows
    ]


def _get_enabled_codes(conn_obj=None) -> list[str]:
    if conn_obj is not None:
        cur = conn_obj.cursor()
        cur.execute("SELECT code FROM product_allowlist WHERE enabled=1 ORDER BY code")
        rows = [_norm_code(r[0]) for r in cur.fetchall()]
        if rows:
            return rows
        return sorted(_get_allow_codes())
    with connect(DATA_DIR) as conn_live:
        return _get_enabled_codes(conn_live)


def _product_allowed(product_key: str, conn_obj=None) -> bool:
    code = _extract_code(product_key)
    if not code:
        return False
    enabled = set(_get_enabled_codes(conn_obj))
    if not enabled:
        return True
    return _norm_code(code) in enabled


def _target_products(conn_obj=None) -> list[tuple[str, str]]:
    defaults = {code: name for code, name in _default_target_products()}
    if conn_obj is not None:
        items = _get_allowlist_from_db(conn_obj)
        enabled = [it for it in items if int(it.get("enabled", 0)) == 1]
        if enabled:
            return [(_norm_code(it["code"]), (it.get("display_name") or defaults.get(_norm_code(it["code"]), _norm_code(it["code"])))) for it in enabled]
        env_codes = sorted(_get_allow_codes())
        if env_codes:
            return [(code, defaults.get(code, code)) for code in env_codes]
        return _default_target_products()
    with connect(DATA_DIR) as conn_live:
        return _target_products(conn_live)


def _pick_row_for_code(code: str, rows: list[dict]) -> dict | None:
    for r in rows:
        pk = str(r.get("product_key", "") or "")
        if pk.startswith(code) or code in pk[:12]:
            return r
    return None


def _fill_target_rows(rows: list[dict], conn_obj=None) -> list[dict]:
    filled = []
    for code, short in _target_products(conn_obj):
        r = _pick_row_for_code(code, rows)
        if not r:
            filled.append({
                "product_key": code,
                "display_key": code,
                "product_name": short,
                "nav": None,
                "unit_nav": None,
                "cum_nav": None,
                "ytd_pct": None,
                "day_pct": None,
                "week_pct": None,
                "month_pct": None,
                "file_ytd_pct": None,
                "year_pct": None,
                "note": "数据缺失",
                "parse_error": "数据缺失",
                "missing": True,
            })
        else:
            rr = dict(r)
            rr["product_name"] = short
            rr.setdefault("note", "")
            rr["missing"] = False
            filled.append(rr)
    return filled


# CODEX_GAPSCAN_HELPERS
def _env_int(key: str, default: int) -> int:
    try:
        return int((os.getenv(key, str(default)) or str(default)).strip())
    except Exception:
        return default


def _env_bool(key: str, default: bool=False) -> bool:
    v = (os.getenv(key, "") or "").strip().lower()
    if v in ("1", "true", "yes", "on"):
        return True
    if v in ("0", "false", "no", "off"):
        return False
    return default


def _list_folders_all_live() -> list[str]:
    from imapclient import IMAPClient
    host = (IMAP_HOST or "").strip()
    user = (IMAP_USER or "").strip()
    pwd = (IMAP_PASS or IMAP_PASSWORD or "").strip()
    if not (host and user and pwd):
        raise RuntimeError("IMAP not configured")
    with IMAPClient(host, ssl=True) as cli:
        cli.login(user, pwd)
        names = [f[2] for f in cli.list_folders()]
    bl = [x.strip().lower() for x in (os.getenv("IMAP_FOLDERS_BLACKLIST", "") or "").split(",") if x.strip()]
    out = []
    for name in names:
        ln = (name or "").lower()
        if any(b in ln for b in bl):
            continue
        out.append(name)
    return out


def _pick_folders(payload: Optional[RunReq]) -> list[str]:
    explicit = [str(x).strip() for x in (getattr(payload, "folders", None) or []) if str(x).strip()]
    if explicit:
        return explicit
    explicit_folder = ""
    if payload is not None and getattr(payload, "folder", None):
        explicit_folder = str(payload.folder).strip()
    if explicit_folder:
        return [explicit_folder]
    if (os.getenv("IMAP_FOLDERS_MODE", "") or "").strip().lower() == "all":
        try:
            folders = _list_folders_all_live()
            if folders:
                return folders
        except Exception as e:
            try:
                _log_fetch("FOLDER_LIST_FAIL", err=str(e))
            except Exception:
                pass
    return [IMAP_FOLDER or "INBOX"]


_CAL = None


def _get_cal():
    global _CAL
    if _CAL is None:
        _CAL = build_calendar_from_env()
        try:
            _CAL.try_update()
        except Exception:
            pass
    return _CAL


def _prev_trading_day_ymd(today: datetime.date) -> str:
    try:
        cal = _get_cal()
        if getattr(cal, "days", None):
            return cal.prev_trading_day(today.strftime("%Y%m%d"))
    except Exception:
        pass
    wd = today.weekday()
    if wd == 0:
        d = today - datetime.timedelta(days=3)
    else:
        d = today - datetime.timedelta(days=1)
    return d.strftime("%Y%m%d")


def _fill_gaps_before(target_ymd: str, max_days: int = 3) -> list[str]:
    filled = []
    try:
        t = datetime.datetime.strptime(target_ymd, "%Y%m%d").date()
    except Exception:
        return filled
    need = []
    d = t
    while len(need) < max_days:
        d = d - datetime.timedelta(days=1)
        if d.weekday() >= 5:
            continue
        need.append(d.strftime("%Y%m%d"))
    for gap in reversed(need):
        with connect(DATA_DIR) as conn_obj:
            cur = conn_obj.cursor()
            cur.execute("SELECT COUNT(*) FROM raw_ingest WHERE date=?", (gap,))
            has_raw = cur.fetchone()[0]
        if has_raw <= 0:
            continue
        try:
            res = process_and_push(RunReq(date_str=gap, force=True, push=False), date_str=gap, force=True)
            if isinstance(res, dict) and res.get("ok"):
                filled.append(gap)
        except Exception:
            continue
    return filled

def _safe_fs_name(name: str, max_len: int = 120) -> str:
    import os
    import re
    import zipfile
    import io

    base = os.path.basename(str(name or "")).replace("/", "_")
    base = re.sub(r"[\r\n\t]+", "_", base)
    base = re.sub(r"\s+", " ", base).strip()
    if not base:
        base = "attachment.xlsx"
    stem, ext = os.path.splitext(base)
    if not ext:
        ext = ".xlsx"
    if len(stem) > max_len:
        stem = stem[:max_len]
    return f"{stem}{ext}"




def _is_workday(d):
    return d.weekday() < 5

def _prev_workday(d):
    x = d
    while True:
        x = x - timedelta(days=1)
        if _is_workday(x):
            return x

def _last_workday_prev_month(d):
    first = d.replace(day=1)
    last_prev = first - timedelta(days=1)
    x = last_prev
    while not _is_workday(x):
        x = x - timedelta(days=1)
    return x

def _prev_week_friday(d):
    monday = d - timedelta(days=d.weekday())
    x = monday - timedelta(days=3)  # last week's Friday
    # safety
    while not _is_workday(x):
        x = x - timedelta(days=1)
    return x


def _clip_utf8(text: str, max_bytes: int) -> str:
    b = (text or "").encode("utf-8", errors="ignore")
    if len(b) <= max_bytes:
        return text or ""
    cut = b[:max_bytes]
    while cut:
        try:
            return cut.decode("utf-8")
        except UnicodeDecodeError:
            cut = cut[:-1]
    return ""



def _is_collect_window(now: datetime.datetime) -> bool:
    # 13:00 <= now <= 16:45
    hhmm = now.strftime("%H%M")
    return "1300" <= hhmm <= "1645"


def _get_state(conn_obj, yyyymmdd: str) -> str:
    cur = conn_obj.cursor()
    cur.execute("SELECT state FROM window_state WHERE date=?", (yyyymmdd,))
    row = cur.fetchone()
    return row[0] if row else "COLLECTING"


def _set_state(conn_obj, yyyymmdd: str, state: str, note: str = "") -> None:
    cur = conn_obj.cursor()
    cur.execute(
        """
        INSERT INTO window_state(date, state, locked_at, note)
        VALUES(?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
          state=excluded.state,
          locked_at=excluded.locked_at,
          note=excluded.note
        """,
        (yyyymmdd, state, _now_iso(), note),
    )
    conn_obj.commit()


def _ensure_aux_tables() -> None:
    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()

        # New install schema
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS raw_ingest (
              date TEXT NOT NULL,
              product_key TEXT NOT NULL,
              raw_path TEXT NOT NULL,
              orig_filename TEXT,
              content_hash TEXT,
              ingest_key TEXT,
              message_id TEXT,
              attachment_idx INTEGER,
              mail_uid TEXT,
              mail_subject TEXT,
              mail_from TEXT,
              received_at TEXT,
              created_at TEXT,
              PRIMARY KEY(date, product_key, raw_path)
            )
            """
        )

        # Runtime migration for existing DBs
        cur.execute("PRAGMA table_info(raw_ingest)")
        cols = [r[1] for r in cur.fetchall()]
        if "ingest_key" not in cols:
            cur.execute("ALTER TABLE raw_ingest ADD COLUMN ingest_key TEXT")
        if "content_hash" not in cols:
            cur.execute("ALTER TABLE raw_ingest ADD COLUMN content_hash TEXT")
        if "message_id" not in cols:
            cur.execute("ALTER TABLE raw_ingest ADD COLUMN message_id TEXT")
        if "attachment_idx" not in cols:
            cur.execute("ALTER TABLE raw_ingest ADD COLUMN attachment_idx INTEGER")

        # Backfill for historical rows without ingest_key
        cur.execute(
            """
            SELECT rowid, date, product_key, raw_path,
                   COALESCE(mail_uid, ""), COALESCE(orig_filename, ""), COALESCE(received_at, "")
            FROM raw_ingest
            WHERE ingest_key IS NULL OR ingest_key = ""
            """
        )
        updates = []
        for row in cur.fetchall():
            rowid, dt, pk, rp, uid, fn, recv = row
            raw = f"{dt}|{pk}|{rp}|{uid}|{fn}|{recv}".encode("utf-8")
            updates.append((_sha256(raw), rowid))
        if updates:
            cur.executemany("UPDATE raw_ingest SET ingest_key=? WHERE rowid=?", updates)

        # Idempotency + query performance
        cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS ux_raw_ingest_date_ingest_key ON raw_ingest(date, ingest_key)"
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_raw_ingest_date_product ON raw_ingest(date, product_key)"
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS window_state (
              date TEXT PRIMARY KEY,
              state TEXT NOT NULL,
              locked_at TEXT,
              note TEXT
            )
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS product_allowlist (
              code TEXT PRIMARY KEY,
              display_name TEXT,
              enabled INTEGER NOT NULL DEFAULT 1,
              updated_at TEXT
            )
            """
        )
        cur.execute("SELECT COUNT(*) FROM product_allowlist")
        if (cur.fetchone() or [0])[0] == 0:
            for code, display_name in _default_target_products():
                cur.execute(
                    "INSERT OR IGNORE INTO product_allowlist(code, display_name, enabled, updated_at) VALUES(?,?,1,datetime('now'))",
                    (code, display_name),
                )
            for code in sorted(_get_allow_codes()):
                cur.execute(
                    "INSERT OR IGNORE INTO product_allowlist(code, display_name, enabled, updated_at) VALUES(?,?,1,datetime('now'))",
                    (code, dict(_default_target_products()).get(code, code)),
                )

        conn_obj.commit()


init_db(DATA_DIR)
_ensure_aux_tables()

TEST_MODE = os.getenv('TEST_MODE','0') == '1'

FETCH_DEBUG = os.getenv("FETCH_DEBUG", "0").strip() in ("1","true","yes","on")
PUBLIC_BASE_URL = os.getenv('PUBLIC_BASE_URL', '').strip()

def _is_test_mode() -> bool:
    return TEST_MODE

def _fmt_pct(x):
    if x is None:
        return "--"
    try:
        return f"{x:+.2f}%"
    except Exception:
        return "--"


# CODEX_CANON_KEY_HELPERS
_CODE_RE = re.compile(r"(SBGZ87|STR134|SXZ218|SXA927)")


def _extract_code(s: str) -> str | None:
    if not s:
        return None
    m = _CODE_RE.search(str(s))
    return m.group(1) if m else None


def _canonical_product_key(product_key: str, product_name: str = "") -> str:
    code = _extract_code(product_key) or _extract_code(product_name)
    return code if code else (product_key or "")


def _series_where(product_key: str) -> tuple[str, tuple]:
    code = _extract_code(product_key)
    if code:
        return "product_key LIKE ?", (code + "%",)
    return "product_key = ?", (product_key,)


def _norm_product_key(x: str) -> str:
    # 归一化产品名：优先归一到产品 code（SBGZ87/STR134/SXZ218/SXA927）
    if x is None:
        return ""
    x = str(x).strip()
    x = x.replace("＿", "_").replace("　", " ")
    x = re.sub(r"\s+", "", x)
    x = re.sub(r"__[^_]{0,64}__uid\d+__.*$", "", x)
    x = re.sub(r"__uid\d+_.*$", "", x)
    x = re.sub(r"_uid\d+_.*$", "", x)
    return _canonical_product_key(x)




def _dedupe_display_rows(rows: list[dict]) -> list[dict]:
    """展示去重：同 norm(product_key) 仅保留 received_at 最新一条。"""
    best = {}
    for r in rows:
        nk = _norm_product_key(r.get("product_key", ""))
        ra = r.get("received_at") or ""
        if nk not in best or ra >= (best[nk].get("received_at") or ""):
            best[nk] = r
    prim = list(best.values())
    prim.sort(key=lambda x: _norm_product_key(x.get("product_key", "")))
    return prim



def _mark_parse_fail(name: str, parse_ok: bool) -> str:
    if not name:
        return name
    return name if parse_ok else f"{name}（解析失败）"

def _fmt_nav_md(v) -> str:
    if v is None:
        return "--"
    try:
        return f"{float(v):.4f}"
    except Exception:
        return "--"


def _fmt_pct_md(v) -> str:
    if v is None:
        return "--"
    try:
        x = float(v)
        # DB里通常存的是比例(0.0123=1.23%)
        if -2.0 <= x <= 2.0:
            x = x * 100.0
        return f"{x:+.2f}%"
    except Exception:
        return "--"


from client_app.app.services.nav_report.utils.render import short_product_name

def _build_wecom_markdown(
    date_str: str,
    rows: list[dict],
    att_cnt: int = 0,
    dup_cnt: int = 0,
    report_url: str = "",
    manifest_url: str = "",
) -> str:
    date_fmt = f"{date_str[0:4]}/{date_str[4:6]}/{date_str[6:8]}" if len(date_str) == 8 else date_str
    lines = []
    lines.append(f"**净值日报 {date_fmt}（测试）**")
    lines.append("")
    lines.append(f"> 产品数：{len(rows)}  附件数：{att_cnt}  重复数：{dup_cnt}")
    lines.append("")

    for r in rows:
        name = r.get("name") or r.get("product_key") or r.get("product_name") or ""
        nav = r.get("nav", "--")
        ytd = r.get("ytd", "--")
        d = r.get("d", "--")
        w = r.get("w", "--")
        m = r.get("m", "--")
        lines.append(f"- **{short_product_name(name, max_len=8)}**")
        lines.append(f"  净值：{nav}  今年来：{ytd}  当日：{d}  当周：{w}  当月：{m}")

    lines.append("")
    lines.append("**下载链接**")
    lines.append(f"- [净值汇总.xlsx]({report_url})")
    lines.append(f"- [原文件清单.xlsx]({manifest_url})")
    return "\n".join(lines)



def _write_raw_manifest(date_str: str, out_dir: str, base_url: str = ""):
    """输出原文件清单（业务可用版）：
    - 只输出必要列（不再生成/隐藏技术列）
    - 重复：同产品(归一化) + 同内容哈希 => dup
      dup 行下沉到后面并整行红底 + 备注
    - 下载列：写入可点击超链接（绝对链接优先）
    """
    from pathlib import Path
    import urllib.parse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    out_dir_path = Path(out_dir)
    out_dir_path.mkdir(parents=True, exist_ok=True)

    abs_base = (base_url or PUBLIC_BASE_URL or "").rstrip("/")

    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()
        cur.execute(
            """
            SELECT
              product_key,
              date,
              orig_filename,
              mail_subject,
              mail_from,
              received_at,
              raw_path,
              content_hash
            FROM raw_ingest
            WHERE date = ?
            ORDER BY received_at ASC
            """,
            (date_str,),
        )
        fetched = cur.fetchall()

    rows = []
    for r in fetched:
        rows.append({
            '备注': (nav_fallback_reason if 'nav_fallback_reason' in locals() else ''),
            "product_key": r[0] or "",
            "date": r[1] or "",
            "orig_filename": r[2] or "",
            "mail_subject": r[3] or "",
            "mail_from": r[4] or "",
            "received_at": r[5] or "",
            "raw_path": r[6] or "",
            "content_hash": (r[7] or "").strip(),
        })

    groups = {}
    for r in rows:
        dup_key = f"{_norm_product_key(r.get('product_key',''))}|{r.get('content_hash','')}"
        groups.setdefault(dup_key, []).append(r)

    first_received = {}
    for k, items in groups.items():
        first_received[k] = min([(x.get("received_at", "") or "") for x in items]) if items else ""

    flat = []
    for dup_key, items in groups.items():
        items_sorted = sorted(items, key=lambda x: (x.get("received_at", "") or ""), reverse=True)
        for idx, r in enumerate(items_sorted):
            rr = dict(r)
            rr["_dup_key"] = dup_key
            rr["_is_dup"] = 1 if idx > 0 else 0
            rr["_first_received"] = first_received.get(dup_key, "")
            flat.append(rr)

    # 排序：产品名 -> 是否重复(0先) -> 收件时间
    flat.sort(key=lambda r: (
        _norm_product_key(r.get("product_key", "")),
        int(r.get("_is_dup", 0)),
        r.get("received_at", "") or "",
    ))

    headers = [
        "产品名称",
        "日期",
        "原文件名",
        "邮件标题",
        "发件人",
        "收件时间",
        "下载原文件",
        "是否重复",
        "重复组",
        "首次收件时间",
        "备注",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "原文件清单"
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    red_fill = PatternFill("solid", fgColor="FFE5E5")
    red_font = Font(color="9C0006")

    link_col = headers.index("下载原文件") + 1

    for r in flat:
        pk = _norm_product_key(r.get("product_key", ""))
        dt = str(r.get("date", ""))
        orig = r.get("orig_filename", "") or ""
        subj = r.get("mail_subject", "") or ""
        mfrom = r.get("mail_from", "") or ""
        recv = r.get("received_at", "") or ""
        raw_path = r.get("raw_path", "") or ""

        # 内容识别：必须包含“基金单位净值/单位净值”等字段，否则跳过
        try:
            ok_field = _file_has_nav_field(str(raw_path))
        except Exception:
            ok_field = False
        if not ok_field:
            print(f"ATT_SKIP no_nav_field: subj_candidate={_subject_is_candidate(subj)} file=unknown")
            continue
        is_dup = int(r.get("_is_dup", 0))
        dup_group = (r.get("content_hash", "") or "")[:8] or (r.get("_dup_key", "")[:8])
        first_recv = r.get("_first_received", "") or ""
        note = f"重复（同组:{dup_group[:8]}，首次:{first_recv}）" if is_dup else ""

        q = urllib.parse.urlencode({"path": raw_path})
        rel = f"/download_raw?{q}"
        url = f"{abs_base}{rel}" if abs_base else rel

        row = [pk, dt, orig, subj, mfrom, recv, "点击下载", is_dup, dup_group, first_recv, note]
        ws.append(row)
        rr = ws.max_row

        link_cell = ws.cell(row=rr, column=link_col)
        link_cell.hyperlink = url
        link_cell.style = "Hyperlink"

        if is_dup == 1:
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=rr, column=c)
                cell.fill = red_fill
                if c != link_col:
                    cell.font = red_font

    for col in range(1, len(headers)+1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row+1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

    out_path = out_dir_path / f"原文件清单-{date_str}.xlsx"
    wb.save(out_path)
    return str(out_path), len(flat)




# === subject/content filter helpers (TEST_SAFE) ===
def _load_csv_env(name: str, default: str = "") -> list[str]:
    s = os.getenv(name, default) or ""
    return [x.strip() for x in s.split(",") if x.strip()]

SUBJECT_HINT_KEYWORDS = _load_csv_env("SUBJECT_HINT_KEYWORDS", "估值表,净值,单位净值,私募证券投资,私募证券投资基金,基金估值,估值,转发")
NAV_FIELD_KEYWORDS = _load_csv_env("NAV_FIELD_KEYWORDS", "基金单位净值,单位净值")

def _subject_is_candidate(subj: str) -> bool:
    if not SUBJECT_HINT_KEYWORDS:
        return True
    s = (subj or "").lower()
    for kw in SUBJECT_HINT_KEYWORDS:
        if kw.lower() in s:
            return True
    return False

def _file_has_nav_field(file_bytes: bytes) -> bool:
    if not NAV_FIELD_KEYWORDS:
        return True
    try:
        blob = file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        blob = ""
    for kw in NAV_FIELD_KEYWORDS:
        if kw and kw in blob:
            return True
    try:
        blob2 = file_bytes.decode("latin1", errors="ignore")
        for kw in NAV_FIELD_KEYWORDS:
            if kw and kw in blob2:
                return True
    except Exception:
        pass
    return False


def _log_fetch(tag: str, **kw):
    try:
        msg = " ".join([f"{k}={v}" for k,v in kw.items()])
        print(f"[FETCH]{tag} {msg}", flush=True)
    except Exception:
        pass




def _imap_day_uids(date_str: str, folder: str) -> list[int]:
    import datetime
    from imapclient import IMAPClient
    d = datetime.datetime.strptime(date_str, "%Y%m%d").date()
    since = d.strftime('%d-%b-%Y')
    before = (d + datetime.timedelta(days=1)).strftime('%d-%b-%Y')

    host = (IMAP_HOST or "").strip()
    user = (IMAP_USER or "").strip()
    pwd = (IMAP_PASS or IMAP_PASSWORD or "").strip()
    if not (host and user and pwd):
        raise RuntimeError('IMAP not configured')

    with IMAPClient(host, ssl=True) as cli:
        cli.login(user, pwd)
        cli.select_folder(folder)
        uids = cli.search(['SINCE', since, 'BEFORE', before])
    return list(uids or [])

app = FastAPI(title="nav-report", version="0.2.0")

# === CLIENT_CONFIG_BRIDGE_BEGIN ===
import json
SERVICE_CONFIG = os.getenv("SERVICE_CONFIG", "").strip()
def _load_client_cfg():
    if not SERVICE_CONFIG or not os.path.exists(SERVICE_CONFIG):
        return None
    try:
        with open(SERVICE_CONFIG, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None
def _apply_client_cfg(cfg: dict):
    imap = (cfg or {}).get("imap", {}) if cfg else {}
    wecom = (cfg or {}).get("wecom", {}) if cfg else {}
    slots = (cfg or {}).get("push_slots", []) if cfg else []
    products = (cfg or {}).get("products", []) if cfg else []
    data_root = (cfg or {}).get("data_root", "") if cfg else ""
    if data_root:
        os.environ["DATA_DIR"] = os.path.join(data_root, "nav_report")
    if imap.get("host"):
        os.environ["IMAP_HOST"] = str(imap.get("host"))
    if imap.get("user"):
        os.environ["IMAP_USER"] = str(imap.get("user"))
    if imap.get("pass"):
        os.environ["IMAP_PASS"] = str(imap.get("pass"))
    if imap.get("lookback_days") is not None:
        os.environ["MAIL_LOOKBACK_DAYS"] = str(imap.get("lookback_days"))
    if imap.get("folder_mode"):
        os.environ["IMAP_FOLDERS_MODE"] = str(imap.get("folder_mode"))
    if imap.get("folders"):
        os.environ["IMAP_FOLDERS"] = ",".join(imap.get("folders"))
    if imap.get("blacklist_keywords"):
        os.environ["IMAP_FOLDERS_BLACKLIST"] = ",".join(imap.get("blacklist_keywords"))
    if wecom.get("webhook_url"):
        os.environ["WECOM_WEBHOOK_URL"] = str(wecom.get("webhook_url"))
    os.environ["WECOM_PUSH_ENABLED"] = "1" if bool(wecom.get("push_enabled", True)) else "0"
    if slots:
        os.environ["PUSH_SLOTS"] = ",".join(slots)
    os.environ["CLIENT_PRODUCTS_JSON"] = json.dumps(products, ensure_ascii=False)
@app.post("/api/reload_config")
def reload_config():
    cfg = _load_client_cfg()
    if cfg:
        _apply_client_cfg(cfg)
    return {"ok": True, "loaded": bool(cfg)}
# === CLIENT_CONFIG_BRIDGE_END ===

app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")



def _explode_zip_attachment(file_bytes: bytes, filename: str):
    """zip -> [(inner_name, inner_bytes)]; non-zip passes through."""
    fn = (filename or '').lower()
    if not fn.endswith('.zip'):
        return [(filename, file_bytes)]

    out = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                inner = info.filename or ''
                base = os.path.basename(inner)
                low = base.lower()
                if not low.endswith(('.xlsx', '.xls', '.xlsm')):
                    continue
                payload = zf.read(info)
                if payload:
                    out.append((base, payload))
    except Exception:
        return [(filename, file_bytes)]
    return out if out else [(filename, file_bytes)]

@app.get("/")
def home():
    return FileResponse(os.path.join(os.path.dirname(__file__), "static", "index.html"))


@app.get("/health")
def health():
    return {"ok": True, "service": "nav-report", "version": "0.2.0"}


@app.get("/api/imap_day_summary")
def imap_day_summary(date_str: str, folder: str = ""):
    fdr = (folder or IMAP_FOLDER or "INBOX").strip() or "INBOX"
    try:
        uids = _imap_day_uids(date_str, fdr)
        return {"ok": True, "date": date_str, "folder": fdr, "total_uids": len(uids), "uids_tail": uids[-20:]}
    except Exception as e:
        return {"ok": False, "date": date_str, "folder": fdr, "error": str(e)}


@app.post("/api/fetch_only")
def fetch_only(request: Request, payload: Optional[RunReq] = None, date_str: str = "", force: bool = False):
    today = _pick_date(payload, date_str)
    force = _pick_force(payload, force)
    lookback_days = _pick_lookback(payload)
    folder = _pick_folder(payload)
    folders = _pick_folders(payload)
    folders = [str(x).strip() for x in folders if str(x).strip()]
    if not folders:
        folders = [IMAP_FOLDER]

    ingest_mode = str(getattr(payload, "ingest_mode", "append") or "append").lower()
    if ingest_mode not in ("append", "replace"):
        ingest_mode = "append"

    recent_only = bool(getattr(payload, "recent_only", False)) if payload is not None else False
    newer_than_hours = int(getattr(payload, "newer_than_hours", 6) or 6) if payload is not None else 6
    since_hours = newer_than_hours if recent_only else max(MAIL_SINCE_HOURS, lookback_days * 24)
    raw_dir, out_dir = yyyymmdd_dir(DATA_DIR, today)

    now = datetime.datetime.now()
    allow_manual_slot = bool(payload is not None and getattr(payload, "slot", "").strip()) and force
    if not _is_collect_window(now):
        # TEST_MODE=1 时允许 force 绕过时间窗；slot 手工入口在生产也允许窗口外补抓
        if not ((TEST_MODE and force) or allow_manual_slot):
            raise HTTPException(status_code=403, detail="outside collect window")

    with connect(DATA_DIR) as conn_obj:
        state = _get_state(conn_obj, today)
        if state != "COLLECTING":
            if not ((_is_test_mode() and force) or allow_manual_slot):
                raise HTTPException(status_code=423, detail=f"collect locked: {state}")

    os.makedirs(raw_dir, exist_ok=True)
    os.makedirs(out_dir, exist_ok=True)

    # STRICT_GUARD: no mails on target date -> short-circuit
    strict = bool(getattr(payload, "strict", True)) if payload is not None else True
    if strict:
        checked = []
        any_uids = False
        for _f in folders:
            try:
                _uids = _imap_day_uids(today, _f)
                checked.append((_f, len(_uids)))
                if _uids:
                    any_uids = True
                    break
            except Exception as e:
                _log_fetch("STRICT_CHECK_FAIL", date=today, folder=_f, err=str(e))
        if checked and not any_uids:
            manifest_path, manifest_rows = _write_raw_manifest(today, out_dir, base_url=(str(request.base_url).rstrip("/") if request is not None else (PUBLIC_BASE_URL or "").strip().rstrip("/")))
            return {
                "ok": True,
                "date": today,
                "attachments": 0,
                "inserted": 0,
                "skipped_dup": 0,
                "replaced": 0,
                "products": 0,
                "strict": True,
                "reason": "NO_MAILS_FOR_DATE",
                "folders": [x[0] for x in checked],
                "imap_uids": {x[0]: x[1] for x in checked},
                "out_manifest": manifest_path,
                "manifest_rows": manifest_rows,
            }

    allowed_ext = (".xlsx", ".xls", ".xlsm", ".zip")

    imap_unseen_only = MAIL_UNSEEN_ONLY and not force

    def _fetch_imap_multi(_folders):
        if not (IMAP_HOST and IMAP_USER and IMAP_PASS):
            raise HTTPException(status_code=400, detail="IMAP not configured")

        all_atts = []
        for _folder in _folders:
            try:
                one = fetch_attachments(
                    host=IMAP_HOST,
                    port=IMAP_PORT,
                    user=IMAP_USER,
                    password=IMAP_PASS,
                    folder=_folder,
                    since_hours=since_hours,
                    unseen_only=imap_unseen_only,
                    from_filter=MAIL_FROM_FILTER,
                    subject_keyword=MAIL_SUBJECT_KEYWORD,
                    subject_regex=MAIL_SUBJECT_REGEX,
                    ext=allowed_ext,
                )
                for a in one:
                    a["_folder"] = _folder
                all_atts.extend(one)
                _log_fetch("FOLDER", folder=_folder, attachments=len(one))
            except Exception as e:
                _log_fetch("FOLDER_SKIP", folder=_folder, err=str(e))
        return all_atts

    atts = []
    if MAIL_PROVIDER == "wecom_mail":
        try:
            mails = wecom_mail.list_inbox_mails(limit=200)
            for m in mails:
                mail_id = m.get("mailid") or m.get("mail_id") or m.get("id")
                if not mail_id:
                    continue
                eml_bytes = wecom_mail.get_mail_raw(str(mail_id))
                headers, raw_atts = wecom_mail.extract_attachments_from_eml(eml_bytes)
                for idx, (orig_filename, attachment_bytes) in enumerate(raw_atts):
                    low = (orig_filename or "").lower()
                    if low and (not low.endswith(allowed_ext)):
                        _log_fetch("EXT_SKIP", uid=str(mail_id), filename=orig_filename)
                        continue
                    atts.append({
                        "filename": orig_filename,
                        "content": attachment_bytes,
                        "uid": str(mail_id),
                        "subject": headers.get("subject") or "",
                        "from": headers.get("from") or "",
                        "received_at": headers.get("date") or _now_iso(),
                        "message_id": headers.get("message_id") or str(mail_id),
                        "attachment_idx": idx,
                        "_folder": "wecom_mail",
                    })
        except Exception as e:
            # fallback to IMAP when wecom_mail credentials are missing/invalid
            if IMAP_HOST and IMAP_USER and IMAP_PASS:
                _log_fetch("WECOM_MAIL_FAIL", err=str(e), action="fallback_imap")
                atts = _fetch_imap_multi(folders)
            else:
                raise HTTPException(status_code=400, detail=f"wecom_mail fetch failed: {e}")
    else:
        atts = _fetch_imap_multi(folders)

    expanded_atts = []
    for a in atts:
        exploded = _explode_zip_attachment(a.get("content", b""), a.get("filename", ""))
        if len(exploded) > 1 or ((a.get("filename") or "").lower().endswith(".zip") and exploded):
            _log_fetch("ZIP_EXPLODE", uid=a.get("uid"), filename=a.get("filename"), exploded=len(exploded))
        base_idx = int(a.get("attachment_idx", 0) or 0)
        for zip_idx, (inner_name, inner_bytes) in enumerate(exploded):
            aa = dict(a)
            aa["filename"] = inner_name or a.get("filename", "")
            aa["content"] = inner_bytes
            aa["attachment_idx"] = base_idx * 1000 + zip_idx
            expanded_atts.append(aa)
    atts = expanded_atts

    _log_fetch("ATT_SUM", date=today, attachments=len(atts), folders=",".join(folders), provider=MAIL_PROVIDER)
    if not atts:
        _log_fetch("NO_ATTACH", date=today, folders=",".join(folders))
        manifest_path, manifest_rows = _write_raw_manifest(today, out_dir, base_url=(str(request.base_url).rstrip("/") if request is not None else (PUBLIC_BASE_URL or "").strip().rstrip("/")))
        return {
            "ok": True,
            "date": today,
            "attachments": 0,
            "inserted": 0,
            "skipped_dup": 0,
            "replaced": 0,
            "products": 0,
            "received_day": today,
            "val_date_buckets": {},
            "out_manifest": manifest_path,
            "manifest_rows": manifest_rows,
        }

    rows = []
    processed_items = []
    skipped_dup = 0
    replaced = 0
    target_val_date_only = bool(getattr(payload, "target_val_date_only", True)) if payload is not None else True
    skipped_valdate_mismatch = 0
    skipped_buckets = {}

    for att_idx, a in enumerate(atts):
        _log_fetch("ATT", uid=a.get("uid"), filename=a.get("filename"), subject=a.get("subject"), folder=a.get("_folder"))
        product_key = _norm_product_key(product_key_from_filename(a["filename"]))
        if not _product_allowed(product_key):
            _log_fetch("SKIP_NOT_ALLOWED", product_key=product_key, uid=a.get("uid"), filename=a.get("filename"))
            continue
        message_id = str(a.get("message_id") or a.get("uid") or "")
        attachment_idx = int(a.get("attachment_idx", att_idx))
        content_hash = hashlib.sha256(a["content"]).hexdigest()
        ingest_key = hashlib.sha256((f"{message_id}|{attachment_idx}|{content_hash}").encode("utf-8")).hexdigest()
        src_folder = a.get("_folder", folder or IMAP_FOLDER)

        ingest_date = _extract_val_date_for_ingest(a["content"], today, a.get("filename", ""))
        processed_items.append({"ingest_date": ingest_date, "uid": a.get("uid"), "filename": a.get("filename"), "orig_filename": a.get("filename")})

        if target_val_date_only and ingest_date and ingest_date != today:
            skipped_valdate_mismatch += 1
            skipped_buckets[ingest_date] = skipped_buckets.get(ingest_date, 0) + 1
            _log_fetch("SKIP_VALDATE_MISMATCH", target=today, got=ingest_date, file=a.get("filename", ""))
            continue

        safe_fn = _safe_fs_name(a["filename"], max_len=36)
        prefix = _clip_utf8(product_key, 48)
        base_name = f"{prefix}__{ingest_key[:10]}__uid{a['uid']}__{safe_fn}"
        save_name = _clip_utf8(base_name, 170)
        if not save_name.lower().endswith(allowed_ext):
            save_name = _clip_utf8(save_name, 165) + ".xlsx"
        raw_path = os.path.join(raw_dir, save_name)

        with connect(DATA_DIR) as conn_obj:
            cur = conn_obj.cursor()
            try:
                logger.info(
                    "[DEDUP_TRACE]exists_check ingest_date=%s target_today=%s mid=%s aidx=%s ch=%s ik=%s fn=%s uid=%s",
                    ingest_date, today, (message_id or "")[:40], attachment_idx, (content_hash or "")[:12], (ingest_key or "")[:12],
                    (a.get("filename") or "")[:80], str(a.get("uid", ""))
                )
            except Exception:
                pass
            cur.execute("SELECT 1 FROM raw_ingest WHERE date=? AND ingest_key=?", (ingest_date, ingest_key))
            exists = cur.fetchone() is not None
            try:
                logger.info(
                    "[DEDUP_TRACE]exists_result ingest_date=%s ik=%s exists=%s uid=%s fn=%s",
                    ingest_date, (ingest_key or "")[:12], exists, str(a.get("uid", "")), (a.get("filename") or "")[:80]
                )
            except Exception:
                pass

        if exists:
            if ingest_mode == "replace":
                with open(raw_path, "wb") as f:
                    f.write(a["content"])
                with connect(DATA_DIR) as conn_obj:
                    cur = conn_obj.cursor()
                    cur.execute(
                        """UPDATE raw_ingest
                           SET product_key=?, raw_path=?, orig_filename=?, content_hash=?, message_id=?, attachment_idx=?,
                               mail_uid=?, mail_subject=?, mail_from=?, received_at=?, created_at=?
                           WHERE date=? AND ingest_key=?""",
                        (
                            product_key,
                            raw_path,
                            a["filename"],
                            content_hash,
                            message_id,
                            attachment_idx,
                            a.get("uid", ""),
                            a.get("subject", ""),
                            a.get("from", ""),
                            a.get("received_at", ""),
                            _now_iso(),
                            ingest_date,
                            ingest_key,
                        ),
                    )
                    conn_obj.commit()
                replaced += 1
                _log_fetch("REPLACE", uid=a.get("uid"), folder=src_folder, filename=a.get("filename"))
            else:
                skipped_dup += 1
                try:
                    logger.info(
                        "[DEDUP_TRACE]SKIP_DUP_HIT ingest_date=%s ik=%s ch=%s mid=%s aidx=%s uid=%s fn=%s",
                        ingest_date, (ingest_key or "")[:12], (content_hash or "")[:12], (message_id or "")[:40],
                        attachment_idx, str(a.get("uid", "")), (a.get("filename") or "")[:80]
                    )
                except Exception:
                    pass
                _log_fetch("SKIP_DUP", uid=a.get("uid"), folder=src_folder, filename=a.get("filename"))
            continue

        with open(raw_path, "wb") as f:
            f.write(a["content"])

        rows.append(
            {
                "product_key": product_key,
                "date": ingest_date,
                "ingest_key": ingest_key,
                "content_hash": content_hash,
                "message_id": message_id,
                "attachment_idx": attachment_idx,
                "raw_path": raw_path,
                "orig_filename": a["filename"],
                "mail_uid": a.get("uid", ""),
                "mail_subject": a.get("subject", ""),
                "mail_from": a.get("from", ""),
                "received_at": a.get("received_at", ""),
            }
        )

    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()
        now_iso = _now_iso()
        for r in rows:
            cur.execute(
                """
                INSERT OR IGNORE INTO raw_ingest
                (date, product_key, raw_path, orig_filename, content_hash, ingest_key, message_id, attachment_idx, mail_uid, mail_subject, mail_from, received_at, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    r.get("date") or today,
                    r["product_key"],
                    r["raw_path"],
                    r.get("orig_filename") or "",
                    r.get("content_hash") or "",
                    r.get("ingest_key") or "",
                    r.get("message_id") or "",
                    r.get("attachment_idx"),
                    r.get("mail_uid") or "",
                    r.get("mail_subject") or "",
                    r.get("mail_from") or "",
                    r.get("received_at") or "",
                    now_iso,
                ),
            )
        conn_obj.commit()

    val_date_buckets = {}
    for _it in processed_items:
        d = (_it.get("ingest_date") or "").strip()
        if d:
            val_date_buckets[d] = val_date_buckets.get(d, 0) + 1

    manifest_path, manifest_rows = _write_raw_manifest(today, out_dir, base_url=(str(request.base_url).rstrip("/") if request is not None else (PUBLIC_BASE_URL or "").strip().rstrip("/")))

    return JSONResponse(
        {
            "ok": True,
            "date": today,
            "attachments": len(atts),
            "inserted": len(rows),
            "skipped_dup": skipped_dup,
            "replaced": replaced,
            "products": len({x["product_key"] for x in rows}),
            "received_day": today,
            "target_val_date_only": target_val_date_only,
            "skipped_valdate_mismatch": skipped_valdate_mismatch,
            "skipped_buckets": skipped_buckets,
            "val_date_buckets": val_date_buckets,
            "out_manifest": manifest_path,
        }
    )


@app.post("/api/fetch_backfill")
def fetch_backfill(request: Request, payload: Optional[RunReq] = None, date_str: str = "", force: bool = False):
    """TEST helper: force backfill outside collect window."""
    if payload is None:
        payload = RunReq()
    payload.force = True
    try:
        return fetch_only(request=request, payload=payload, date_str=date_str, force=True)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"backfill failed: {e}")



def _prev_weekday_yyyymmdd(d: str) -> str:
    import datetime
    dt = datetime.datetime.strptime(d, "%Y%m%d").date() - datetime.timedelta(days=1)
    while dt.weekday() >= 5:
        dt -= datetime.timedelta(days=1)
    return dt.strftime("%Y%m%d")


def _get_prev_date_from_series(conn, product_key: str, target_date: str):
    where_sql, where_args = _series_where(product_key)
    row = conn.execute(
        f"SELECT date FROM nav_series WHERE {where_sql} AND date < ? ORDER BY date DESC LIMIT 1",
        (*where_args, target_date),
    ).fetchone()
    return row[0] if row else None


def _has_series_point(conn, product_key: str, d: str) -> bool:
    where_sql, where_args = _series_where(product_key)
    return conn.execute(
        f"SELECT 1 FROM nav_series WHERE {where_sql} AND date=? LIMIT 1",
        (*where_args, d),
    ).fetchone() is not None


def _autofill_prev_day_from_raw_ingest(conn, target_date: str):
    """TEST helper: fill nav_series(prev_date) from raw_ingest(prev_date) when missing."""
    # target products from today's raw_ingest first (more direct than nav_series)
    pks = [r[0] for r in conn.execute("SELECT DISTINCT product_key FROM raw_ingest WHERE date=?", (target_date,)).fetchall()]
    if not pks:
        pks = [r[0] for r in conn.execute("SELECT DISTINCT product_key FROM nav_series WHERE date=?", (target_date,)).fetchall()]
    if not pks:
        return {"ok": True, "target": target_date, "prev_date": None, "need_fill": 0, "filled": 0, "skipped_no_raw": 0, "errors": []}

    prev_candidates = [d for d in (_get_prev_date_from_series(conn, pk, target_date) for pk in pks) if d]
    prev_date = sorted(prev_candidates)[-1] if prev_candidates else _prev_weekday_yyyymmdd(target_date)

    need = [pk for pk in pks if not _has_series_point(conn, pk, prev_date)]
    if not need:
        return {"ok": True, "target": target_date, "prev_date": prev_date, "need_fill": 0, "filled": 0, "skipped_no_raw": 0, "errors": []}

    filled = 0
    skipped_no_raw = 0
    errors = []

    for pk in need:
        row = conn.execute(
            "SELECT raw_path, orig_filename FROM raw_ingest WHERE date=? AND product_key=? ORDER BY received_at DESC, created_at DESC LIMIT 1",
            (prev_date, pk),
        ).fetchone()
        if not row:
            skipped_no_raw += 1
            continue
        raw_path, orig_filename = row
        try:
            b = Path(raw_path).read_bytes()
        except Exception as e:
            errors.append(f"{pk}:read_raw_fail:{e}")
            continue

        try:
            nav = None
            if '_extract_nav_any' in globals():
                nav, _reason = _extract_nav_any(b)
            m = parse_metrics(b) or {}
            unit_nav = m.get('unit_nav')
            cum_nav = m.get('cum_nav')
            if nav is None:
                nav = unit_nav if unit_nav is not None else m.get('nav')
            if nav is None:
                errors.append(f"{pk}:parse_nav_fail")
                continue
            upsert_nav(conn, pk, prev_date, nav=nav, unit_nav=unit_nav, cum_nav=cum_nav)
            filled += 1
        except Exception as e:
            errors.append(f"{pk}:upsert_fail:{e}")

    _log_fetch("AUTO_PREV", target=target_date, prev=prev_date, need=len(need), filled=filled, no_raw=skipped_no_raw, err=len(errors))
    return {"ok": True, "target": target_date, "prev_date": prev_date, "need_fill": len(need), "filled": filled, "skipped_no_raw": skipped_no_raw, "errors": errors}


@app.post("/api/process_and_push")
def process_and_push(payload: Optional[RunReq] = None, date_str: str = "", force: bool = False):
    from client_app.app.services.nav_report.utils.wecom import get_token, upload_media, send_chat_text, send_chat_image, send_chat_file
    from client_app.app.services.nav_report.utils.render import build_excel, render_table_image, fmt_pct, fmt_num, short_product_name

    today = _pick_date(payload, date_str)
    force = _pick_force(payload, force)
    _raw_dir, out_dir = yyyymmdd_dir(DATA_DIR, today)
    os.makedirs(out_dir, exist_ok=True)

    with connect(DATA_DIR) as conn_obj:
        state = _get_state(conn_obj, today)
        if state == "DONE" and not force:
            return {"ok": True, "date": today, "status": "already_done"}
        if state not in ("COLLECTING", "DONE"):
            raise HTTPException(status_code=423, detail=f"processing locked: {state}")
        _set_state(conn_obj, today, "PROCESSING", note="batch start")

    try:
        with connect(DATA_DIR) as conn_obj:
            cur = conn_obj.cursor()
            cur.execute(
                """
                SELECT product_key, raw_path, orig_filename, mail_uid, mail_subject, mail_from, received_at, created_at
                FROM raw_ingest
                WHERE date=?
                """,
                (today,),
            )
            items = cur.fetchall()


        # AUTO_PREV_FILL_HOOK (TEST only): best-effort fill previous day baselines
        try:
            if os.getenv("TEST_MODE", "0") == "1":
                with connect(DATA_DIR) as conn_fill:
                    _autofill_prev_day_from_raw_ingest(conn_fill, today)
        except Exception as _e:
            _log_fetch("AUTO_PREV_FAIL", err=str(_e))

        if not items:
            with connect(DATA_DIR) as conn_obj:
                _set_state(conn_obj, today, "COLLECTING", note="no raw_ingest for today")
            return {"ok": True, "skipped": True, "reason": "NO_DATA_FOR_DATE: raw_ingest empty, please fetch_only first", "date": today, "products": 0}

        def is_newer(a, b):
            if b is None:
                return True
            ka = (a[5] or "", a[2] or "", a[6] or "")
            kb = (b[5] or "", b[2] or "", b[6] or "")
            return ka > kb

        latest = {}
        for raw_product_key, raw_path, orig_filename, mail_uid, mail_subject, mail_from, received_at, created_at in items:
            product_key = _norm_product_key(raw_product_key)
            cand = (raw_path, orig_filename, mail_uid, mail_subject, mail_from, received_at, created_at)
            prev = latest.get(product_key)
            if is_newer(cand, prev):
                latest[product_key] = cand

        rows = []
        history_upserts = 0
        for product_key, (raw_path, _orig_filename, mail_uid, mail_subject, mail_from, received_at, _created_at) in latest.items():
            b = Path(raw_path).read_bytes()
            try:
                history_upserts += _import_nav_history_from_bytes(b, fallback_product_key=product_key)
            except Exception as e:
                _log_fetch("NAV_HISTORY_IMPORT_FAIL", product_key=product_key, err=str(e))
            parse_ok = True
            try:
                m = parse_metrics(b) or {}
            except Exception:
                m = {}
                parse_ok = False
            unit_nav = m.get("unit_nav")
            cum_nav = m.get("cum_nav")
            nav_val = unit_nav if unit_nav not in (None, "") else m.get("nav")
            parse_error = ""
            val_date = today

            if nav_val in (None, ""):
                fb_nav, fb_reason = _extract_nav_any(b)
                if fb_nav is not None:
                    nav_val = fb_nav
                    if unit_nav in (None, ""):
                        unit_nav = fb_nav
                    parse_ok = True
                    parse_error = f"fallback:{fb_reason}"
                    _log_fetch("NAV_FALLBACK_OK", product_key=product_key, reason=fb_reason)
                else:
                    parse_ok = False
                    parse_error = fb_reason or "nav_not_found"
                    _log_fetch("NAV_FALLBACK_FAIL", product_key=product_key, reason=parse_error)

            if nav_val not in (None, ""):
                val_date = today  # force nav_series date from raw_ingest/process target date
                try:
                    with connect(DATA_DIR) as conn_series:
                        upsert_nav(conn_series, product_key, val_date, nav=nav_val, unit_nav=unit_nav, cum_nav=cum_nav)
                        s_day, s_week, s_month, s_ytd = _calc_returns_from_series(conn_series, product_key, val_date)
                    m["day_pct"] = s_day
                    m["week_pct"] = s_week
                    m["month_pct"] = s_month
                    m["year_pct"] = s_ytd
                except Exception as e:
                    _log_fetch("SERIES_CALC_FAIL", product_key=product_key, err=str(e))

            rows.append(
                {
                    "product_key": product_key,
                    "display_key": _mark_parse_fail(product_key, parse_ok),
                    "parse_ok": parse_ok,
                    "parse_error": parse_error,
                    "note": parse_error,
                    "date": today,
                    "nav": nav_val,
                    "unit_nav": unit_nav,
                    "cum_nav": cum_nav,
                    "day_pct": m.get("day_pct"),
                    "week_pct": m.get("week_pct"),
                    "month_pct": m.get("month_pct"),
                    "file_ytd_pct": m.get("file_ytd_pct"),
                    "year_pct": m.get("year_pct"),
                    "raw_path": raw_path,
                    "mail_uid": mail_uid,
                    "mail_subject": mail_subject,
                    "mail_from": mail_from,
                    "received_at": received_at,
                }
            )

        def fetch_nav(conn_obj, product_key: str, ymd: str):
            cur = conn_obj.cursor()
            cur.execute(
                "SELECT nav FROM daily_metrics WHERE product_key=? AND date=? AND nav IS NOT NULL",
                (product_key, ymd),
            )
            row = cur.fetchone()
            return row[0] if row else None

        def find_year_base(conn_obj, product_key: str, ymd: str, mode: str):
            jan1 = f"{ymd[:4]}0101"
            cur = conn_obj.cursor()
            if mode == "STRICT":
                return jan1, fetch_nav(conn_obj, product_key, jan1)
            cur.execute(
                """
                SELECT date, nav
                FROM daily_metrics
                WHERE product_key=? AND date>=? AND date<=? AND nav IS NOT NULL
                ORDER BY date ASC
                LIMIT 1
                """,
                (product_key, jan1, ymd),
            )
            row = cur.fetchone()
            return (row[0], row[1]) if row else (None, None)

        mode = YTD_BASE_MODE

        with connect(DATA_DIR) as conn_obj:
            cur = conn_obj.cursor()
            for r in rows:
                cur.execute(
                    """
                    INSERT INTO daily_metrics
                    (product_key, date, nav, day_pct, week_pct, month_pct, ytd_pct, ytd_mode,
                     source_nav, source_day, source_week, source_month,
                     raw_path, mail_uid, mail_subject, mail_from, received_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(product_key, date) DO UPDATE SET
                      nav=excluded.nav,
                      day_pct=excluded.day_pct,
                      week_pct=excluded.week_pct,
                      month_pct=excluded.month_pct,
                      ytd_pct=excluded.ytd_pct,
                      raw_path=excluded.raw_path,
                      mail_uid=excluded.mail_uid,
                      mail_subject=excluded.mail_subject,
                      mail_from=excluded.mail_from,
                      received_at=excluded.received_at
                    """,
                    (
                        r.get("display_key") or r["product_key"],
                        r["date"],
                        r["nav"],
                        r["day_pct"],
                        r["week_pct"],
                        r["month_pct"],
                        None,
                        mode,
                        "file",
                        "file",
                        "file",
                        "file",
                        r["raw_path"],
                        r["mail_uid"],
                        r["mail_subject"],
                        r["mail_from"],
                        r["received_at"],
                    ),
                )

            for r in rows:
                file_ytd_raw = r.get("year_pct") if r.get("year_pct") is not None else r.get("file_ytd_pct")
                file_ytd = _to_ratio_pct_input(file_ytd_raw)
                if file_ytd is not None:
                    cur.execute(
                        "UPDATE daily_metrics SET ytd_pct=? WHERE product_key=? AND date=?",
                        (file_ytd, r["product_key"], today),
                    )
                    continue

                nav_today = r.get("nav")
                if not nav_today:
                    continue

                _base_date, base_nav = find_year_base(conn_obj, r["product_key"], today, mode)
                if not base_nav:
                    continue

                ytd = (nav_today - base_nav) / base_nav
                cur.execute(
                    "UPDATE daily_metrics SET ytd_pct=? WHERE product_key=? AND date=?",
                    (ytd, r["product_key"], today),
                )

            conn_obj.commit()

            for r in rows:
                cur.execute(
                    "SELECT ytd_pct FROM daily_metrics WHERE product_key=? AND date=?",
                    (r["product_key"], today),
                )
                row = cur.fetchone()
                r["ytd_pct"] = row[0] if row else None

        rows = _dedupe_display_rows(rows)
        rows = [r for r in rows if _product_allowed(r.get("product_key", ""))]
        rows = _fill_target_rows(rows)

        dt = datetime.datetime.strptime(today, "%Y%m%d")
        nav_col = f"{dt.month}.{dt.day}净值"

        out_xlsx = os.path.join(out_dir, f"净值汇总-{today}.xlsx")
        out_manifest = os.path.join(out_dir, f"原文件清单-{today}.xlsx")
        # FINAL_SERIES_OVERRIDE: enforce returns from nav_series only
        try:
            with connect(DATA_DIR) as conn_series:
                for r in rows:
                    pk = r.get("product_key")
                    ymd = r.get("val_date") or today
                    d, w, mn, y = _calc_returns_from_series(conn_series, pk, ymd)
                    r["day_pct"] = _calc_day_from_unit_series(conn_series, pk, ymd)
                    r["week_pct"] = w
                    r["month_pct"] = mn
                    r["year_pct"] = y
                    r["file_ytd_pct"] = y
                    r["ytd_pct"] = y
        except Exception as e:
            _log_fetch("FINAL_SERIES_OVERRIDE_FAIL", err=str(e))

        out_img_today = os.path.join(out_dir, f"净值日报-{today}.png")
        stale_out_img_cmp = os.path.join(out_dir, f"昨日对比-{today}.png")
        if os.path.exists(stale_out_img_cmp):
            try:
                os.remove(stale_out_img_cmp)
            except Exception:
                pass
        out_img_cmp = None

        build_excel(out_xlsx, today, rows)
        try:
            _write_raw_manifest_from_db(today, out_manifest, base_url=(PUBLIC_BASE_URL or "").strip())
            _log_fetch("MANIFEST_REWRITE", date=today, path=out_manifest)
        except Exception as e:
            _log_fetch("MANIFEST_REWRITE_FAIL", date=today, err=str(e))

        title1 = f"净值日报 {dt.year}/{dt.month:02d}/{dt.day:02d}"
        headers1 = ["产品名称", nav_col, "今年来", "当日", "当周", "当月"]
        img1_rows = []
        for r in rows:
            img1_rows.append(
                [
                    short_product_name(r["product_key"], max_len=8),
                    fmt_num(r.get("nav"), 4),
                    fmt_pct(r.get("ytd_pct")),
                    fmt_pct(r.get("day_pct")),
                    fmt_pct(r.get("week_pct")),
                    fmt_pct(r.get("month_pct")),
                ]
            )
        render_table_image(out_img_today, title1, headers1, img1_rows, highlight_cols=[2, 3, 4, 5])

        # disabled: no 昨日对比 image

        corp_id = os.getenv("WECOM_CORP_ID", "").strip()
        agent_id = os.getenv("WECOM_AGENT_ID", "").strip()
        secret = os.getenv("WECOM_APP_SECRET", "").strip()
        chatid = os.getenv("WECOM_CHAT_ID", "").strip()
        push_allowed = not (payload is not None and getattr(payload, "push", True) is False)

        pushed = False
        push_error = ""
        if push_allowed and corp_id and agent_id and secret and chatid:
            try:
                token = get_token(corp_id, secret)
                send_chat_text(token, chatid, f"净值批处理完成：{today}\\n产品数：{len(rows)}")

                mid_img1 = upload_media(token, "image", os.path.basename(out_img_today), Path(out_img_today).read_bytes())
                send_chat_image(token, chatid, mid_img1)

                # disabled: no 昨日对比 image

                mid_file = upload_media(token, "file", os.path.basename(out_xlsx), Path(out_xlsx).read_bytes())
                send_chat_file(token, chatid, mid_file)
                pushed = True
            except Exception as e:
                push_error = str(e)

        # --- WEBHOOK_FALLBACK_SEND_PROD ---
        try:
            push_enabled = os.getenv("WECOM_PUSH_ENABLED", "0").strip() == "1"
            webhook_url_or_key = (os.getenv("WECOM_WEBHOOK_URL", "").strip() or os.getenv("WECOM_WEBHOOK_KEY", "").strip())
            if push_allowed and (not pushed) and push_enabled and webhook_url_or_key:
                import base64
                from urllib.parse import quote
                from client_app.app.services.nav_report.utils.wecom_webhook import send_markdown, send_image_b64, send_file_b64

                rows_for_md = []
                att_cnt = 0
                dup_cnt = 0
                with connect(DATA_DIR) as conn_obj:
                    cur = conn_obj.cursor()
                    cur.execute(
                        """
                        SELECT product_key, nav, ytd_pct, day_pct, week_pct, month_pct
                        FROM daily_metrics
                        WHERE date=?
                        ORDER BY product_key ASC
                        """,
                        (today,),
                    )
                    for name, nav, ytd, d, w, m in cur.fetchall():
                        rows_for_md.append({
                            "name": name or "",
                            "nav": _fmt_nav_md(nav),
                            "ytd": _fmt_pct_md(ytd),
                            "d": _fmt_pct_md(d),
                            "w": _fmt_pct_md(w),
                            "m": _fmt_pct_md(m),
                        })
                    cur.execute("SELECT COUNT(*) FROM raw_ingest WHERE date=?", (today,))
                    one = cur.fetchone()
                    att_cnt = int(one[0] or 0) if one else 0
                    cur.execute(
                        """
                        SELECT COALESCE(SUM(cnt - 1), 0) FROM (
                            SELECT COALESCE(product_key,''), COALESCE(content_hash,''), COUNT(*) AS cnt
                            FROM raw_ingest
                            WHERE date=?
                            GROUP BY COALESCE(product_key,''), COALESCE(content_hash,'')
                            HAVING COUNT(*) > 1
                        ) t
                        """,
                        (today,),
                    )
                    one = cur.fetchone()
                    dup_cnt = int(one[0] or 0) if one else 0

                base_url = (PUBLIC_BASE_URL or "").strip().rstrip("/")
                if not base_url:
                    base_url = "http://106.55.177.194:6003"
                report_url = f"{base_url}/download?path={quote(out_xlsx, safe='')}"
                manifest_url = f"{base_url}/download?path={quote(out_manifest, safe='')}"
                markdown_content = _build_wecom_markdown(
                    today,
                    rows_for_md,
                    att_cnt=att_cnt,
                    dup_cnt=dup_cnt,
                    report_url=report_url,
                    manifest_url=manifest_url,
                )
                send_markdown(webhook_url_or_key, markdown_content)
                with open(out_img_today, "rb") as f:
                    send_image_b64(webhook_url_or_key, base64.b64encode(f.read()).decode("utf-8"))
                # disabled: no 昨日对比 image
                with open(out_xlsx, "rb") as f:
                    send_file_b64(webhook_url_or_key, os.path.basename(out_xlsx), base64.b64encode(f.read()).decode("utf-8"))
                pushed = True
                push_error = ""
        except Exception as e:
            if not push_error:
                push_error = f"webhook_push_fail:{e}"
        # --- WEBHOOK_FALLBACK_SEND_PROD end ---

        with connect(DATA_DIR) as conn_obj:
            _set_state(conn_obj, today, "DONE", note="batch done")

        return {
            "ok": True,
            "date": today,
            "products": len(rows),
            "history_upserts": history_upserts,
            "out_dir": out_dir,
            "out_xlsx": out_xlsx,
            "out_manifest": out_manifest,
            "out_img_today": out_img_today,
            "pushed": pushed,
            "push_error": push_error,
        }

    except HTTPException:
        raise
    except Exception as e:
        with connect(DATA_DIR) as conn_obj:
            _set_state(conn_obj, today, "COLLECTING", note=f"batch failed: {e}")
        raise



if TEST_MODE:
    @app.post("/api/upload_test")
    def upload_test(files: list[UploadFile] = File(...), date_str: str = Form("")):
        """
        仅用于网页版测试：上传 xlsx -> 写入 raw_ingest + 落盘 raw（不走 IMAP，不受收集窗口限制）
        """
        import datetime, os
        from pathlib import Path
        from client_app.app.services.nav_report.utils.naming import product_key_from_filename, yyyymmdd_dir
        from client_app.app.services.nav_report.utils.db import connect

        today = (date_str or "").strip()
        if len(today) != 8:
            today = datetime.datetime.now().strftime("%Y%m%d")

        raw_dir, out_dir = yyyymmdd_dir(DATA_DIR, today)
        os.makedirs(raw_dir, exist_ok=True)
        os.makedirs(out_dir, exist_ok=True)

        inserted = 0
        skipped = 0
        items = []

        with connect(DATA_DIR) as conn:
            cur = conn.cursor()
            for f in files:
                if not (f.filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls") or filename.lower().endswith(".xlsm") or filename.lower().endswith(".zip") or f.filename.lower().endswith(".xls") or f.filename.lower().endswith(".csv")):
                    skipped += 1
                    continue
                content = f.file.read()
                uid = "WEBTEST"
                ingest_key = _make_ingest_key(uid, f.filename, content)
                # 幂等：同一天同 ingest_key 不重复
                cur.execute("SELECT 1 FROM raw_ingest WHERE date=? AND ingest_key=?", (today, ingest_key))
                if cur.fetchone():
                    skipped += 1
                    continue

                product_key = _norm_product_key(product_key_from_filename(f.filename))
                safe_fn = _safe_fs_name(f.filename, max_len=36)
                prefix = _clip_utf8(product_key, 48)
                base_name = f"{prefix}__{ingest_key[:10]}__uid{uid}__{safe_fn}"
                save_name = _clip_utf8(base_name, 170)
                if not save_name.lower().endswith(".xlsx") or filename.lower().endswith(".xls") or filename.lower().endswith(".xlsm") or filename.lower().endswith(".zip"):
                    save_name = _clip_utf8(save_name, 165) + ".xlsx"
                raw_path = os.path.join(raw_dir, save_name)
                Path(raw_path).write_bytes(content)

                cur.execute("""
                INSERT OR IGNORE INTO raw_ingest
                (date, product_key, raw_path, orig_filename, ingest_key, mail_uid, mail_subject, mail_from, received_at, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (
                    today, product_key, raw_path, f.filename, ingest_key, uid,
                    "WEB TEST", "web@test", _now_iso(), _now_iso()
                ))
                inserted += 1
                items.append({"product_key": product_key, "raw_path": raw_path, "ingest_key": ingest_key})

            conn.commit()

        return {"ok": True, "date": today, "inserted": inserted, "skipped": skipped, "items": items}


if TEST_MODE:
    @app.get("/api/artifacts")
    def artifacts(date_str: str = ""):
        import datetime, os
        from client_app.app.services.nav_report.utils.naming import yyyymmdd_dir

        today = (date_str or "").strip()
        if len(today) != 8:
            today = datetime.datetime.now().strftime("%Y%m%d")

        _, out_dir = yyyymmdd_dir(DATA_DIR, today)
        if not os.path.isdir(out_dir):
            return {"ok": True, "date": today, "out_dir": out_dir, "items": [], "files": []}

        items = []
        files = []
        for name in sorted(os.listdir(out_dir)):
            if not name.lower().endswith((".xlsx", ".png", ".csv")):
                continue
            fp = os.path.join(out_dir, name)
            if not os.path.isfile(fp):
                continue
            files.append(name)
            items.append({"name": name, "path": fp, "size": os.path.getsize(fp)})

        return {"ok": True, "date": today, "out_dir": out_dir, "items": items, "files": files}




if TEST_MODE:
    @app.api_route("/download_raw", methods=["GET","HEAD"])
    def download_raw(path: str = Query(..., description="absolute path under DATA_DIR/raw only")):
        """
        安全下载：仅允许 DATA_DIR/raw 下文件。
        """
        import os

        base = os.path.abspath(os.environ.get("DATA_DIR", "data_test"))
        raw_base = os.path.abspath(os.path.join(base, "raw"))
        ap = os.path.abspath(path)

        if not ap.startswith(raw_base + os.sep):
            raise HTTPException(status_code=403, detail="forbidden path")
        if not os.path.isfile(ap):
            raise HTTPException(status_code=404, detail="file not found")

        return FileResponse(ap, filename=os.path.basename(ap), media_type="application/octet-stream")

if TEST_MODE:
    @app.get("/download")
    def download(path: str):
        """
        安全下载：仅允许 DATA_DIR/out 下文件。
        """
        import os

        base = os.path.abspath(os.environ.get("DATA_DIR", "data_test"))
        out_base = os.path.abspath(os.path.join(base, "out"))
        ap = os.path.abspath(path)

        if not ap.startswith(out_base + os.sep):
            raise HTTPException(status_code=403, detail="forbidden path")
        if not os.path.isfile(ap):
            raise HTTPException(status_code=404, detail="file not found")

        return FileResponse(ap, filename=os.path.basename(ap), media_type="application/octet-stream")


if TEST_MODE:
    @app.post("/api/push_today")
    def push_today(req: Optional[RunReq] = None):
        """
        只推送今日已生成产物（Excel + 2张图），不重新生成。
        同时发送 Markdown 日报摘要（产品列表 + 下载链接）。
        """
        import os
        import base64
        from urllib.parse import quote
        from client_app.app.services.nav_report.utils.wecom_webhook import send_markdown, send_image_b64, send_file_b64

        if os.environ.get("WECOM_PUSH_ENABLED", "0") != "1":
            raise HTTPException(status_code=400, detail="WECOM_PUSH_DISABLED")

        webhook_url = os.environ.get("WECOM_WEBHOOK_URL", "").strip()
        webhook_url_or_key = (os.environ.get("WECOM_WEBHOOK_URL","").strip() or os.environ.get("WECOM_WEBHOOK_KEY","").strip())
        if not webhook_url_or_key:
            raise HTTPException(status_code=400, detail="WECOM_WEBHOOK_MISSING")

        date_str = _pick_date(req, "")
        base = os.path.abspath(os.environ.get("DATA_DIR", "data_test"))
        out_dir = os.path.join(base, "out", date_str[:4], date_str[4:6], date_str[6:8])

        if not os.path.isdir(out_dir):
            raise HTTPException(status_code=404, detail=f"out dir not found: {out_dir}")

        xlsx = os.path.join(out_dir, f"净值汇总-{date_str}.xlsx")
        manifest_xlsx = os.path.join(out_dir, f"原文件清单-{date_str}.xlsx")
        img1 = os.path.join(out_dir, f"净值日报-{date_str}.png")

        for fp in (xlsx, img1):
            if not os.path.exists(fp):
                raise HTTPException(status_code=404, detail=f"missing artifact: {os.path.basename(fp)}")

        # 汇总产品列表 + 附件统计 + 重复统计
        rows_for_md = []
        att_cnt = 0
        dup_cnt = 0
        with connect(DATA_DIR) as conn_obj:
            cur = conn_obj.cursor()
            cur.execute(
                """
                SELECT product_key, nav, ytd_pct, day_pct, week_pct, month_pct
                FROM daily_metrics
                WHERE date=?
                ORDER BY product_key ASC
                """,
                (date_str,),
            )
            for name, nav, ytd, d, w, m in cur.fetchall():
                rows_for_md.append({
                    "name": name or "",
                    "nav": _fmt_nav_md(nav),
                    "ytd": _fmt_pct_md(ytd),
                    "d": _fmt_pct_md(d),
                    "w": _fmt_pct_md(w),
                    "m": _fmt_pct_md(m),
                })

            cur.execute("SELECT COUNT(*) FROM raw_ingest WHERE date=?", (date_str,))
            one = cur.fetchone()
            att_cnt = int(one[0] or 0) if one else 0

            cur.execute(
                """
                SELECT COALESCE(SUM(cnt - 1), 0) FROM (
                    SELECT COALESCE(product_key,''), COALESCE(content_hash,''), COUNT(*) AS cnt
                    FROM raw_ingest
                    WHERE date=?
                    GROUP BY COALESCE(product_key,''), COALESCE(content_hash,'')
                    HAVING COUNT(*) > 1
                ) t
                """,
                (date_str,),
            )
            one = cur.fetchone()
            dup_cnt = int(one[0] or 0) if one else 0

        base_url = (PUBLIC_BASE_URL or "").strip().rstrip("/")
        if not base_url:
            base_url = "http://106.55.177.194:6003"

        report_url = f"{base_url}/download?path={quote(xlsx, safe='')}"
        manifest_url = f"{base_url}/download?path={quote(manifest_xlsx, safe='')}"

        markdown_content = _build_wecom_markdown(
            date_str,
            rows_for_md,
            att_cnt=att_cnt,
            dup_cnt=dup_cnt,
            report_url=report_url,
            manifest_url=manifest_url,
        )

        send_markdown(webhook_url_or_key, markdown_content)

        with open(img1, "rb") as f:
            send_image_b64(webhook_url_or_key, base64.b64encode(f.read()).decode("utf-8"))
        # disabled: no 昨日对比 image
        with open(xlsx, "rb") as f:
            send_file_b64(webhook_url_or_key, os.path.basename(xlsx), base64.b64encode(f.read()).decode("utf-8"))

        return {"ok": True, "date": date_str, "sent": ["markdown", "img1", "xlsx"]}



if TEST_MODE:
    @app.post("/api/reset_test_day")
    def reset_test_day(payload: Optional[RunReq] = None, date_str: str = ""):
        """
        测试专用：清空当日测试数据（raw_ingest + raw 文件 + out 产物）
        不影响其他日期；生产模式不要调用。
        """
        import datetime, os, shutil
        from client_app.app.services.nav_report.utils.db import connect
        from client_app.app.services.nav_report.utils.naming import yyyymmdd_dir

        today = _pick_date(payload, date_str)

        raw_dir, out_dir = yyyymmdd_dir(DATA_DIR, today)

        # 1) 清 DB：当天 raw_ingest + daily_metrics（避免上一轮 ytd/nav 影响）
        with connect(DATA_DIR) as conn:
            cur = conn.cursor()
            # raw_ingest 清理
            cur.execute("DELETE FROM raw_ingest WHERE date=?", (today,))
            # daily_metrics 可能存在就清（表不存在也别炸）
            try:
                cur.execute("DELETE FROM daily_metrics WHERE date=?", (today,))
            except Exception:
                pass
            # window_state 清理（让测试回到 COLLECTING）
            try:
                cur.execute("DELETE FROM window_state WHERE date=?", (today,))
            except Exception:
                pass
            conn.commit()

        # 2) 清目录：只删当日目录
        def safe_rmtree(d):
            if os.path.isdir(d) and os.path.realpath(d).startswith(os.path.realpath(DATA_DIR)):
                shutil.rmtree(d, ignore_errors=True)

        safe_rmtree(raw_dir)
        safe_rmtree(out_dir)

        return {"ok": True, "date": today, "cleared": True, "raw_dir": raw_dir, "out_dir": out_dir}


def _calc_returns_by_nav(conn, product_key: str, valuation_ymd: str):
    """
    returns: (ytd, day, week, month) as floats, None if base missing
    YTD: A口径=本年首个可用净值（>= Jan 1）
    """
    import datetime as _dt
    T = _dt.datetime.strptime(valuation_ymd, "%Y%m%d").date()

    def get_nav(ymd: str):
        row = conn.execute("SELECT nav FROM nav_series WHERE product_key=? AND date=?",
                           (product_key, ymd)).fetchone()
        return None if not row else float(row[0])

    # day
    prev_d = _prev_workday(T)
    nav_t = get_nav(valuation_ymd)
    nav_prev = get_nav(prev_d.strftime("%Y%m%d"))

    day = (nav_t / nav_prev - 1.0) if (nav_t and nav_prev) else None

    # week
    w_base = _prev_week_friday(T)
    nav_w = get_nav(w_base.strftime("%Y%m%d"))
    week = (nav_t / nav_w - 1.0) if (nav_t and nav_w) else None

    # month
    m_base = _last_workday_prev_month(T)
    nav_m = get_nav(m_base.strftime("%Y%m%d"))
    month = (nav_t / nav_m - 1.0) if (nav_t and nav_m) else None

    # ytd (A): first available in year
    year_start = _dt.date(T.year, 1, 1).strftime("%Y%m%d")
    row = conn.execute("""
        SELECT nav FROM nav_series
        WHERE product_key=? AND date>=? AND date<=?
        ORDER BY date ASC LIMIT 1
    """, (product_key, year_start, valuation_ymd)).fetchone()
    nav_y0 = None if not row else float(row[0])
    ytd = (nav_t / nav_y0 - 1.0) if (nav_t and nav_y0) else None

    return (ytd, day, week, month)




@app.post("/api/process_prev_trading")
def api_process_prev_trading(payload: Optional[RunReq] = None):
    payload = payload or RunReq()
    D = _prev_trading_day_ymd(datetime.date.today())
    max_days = _env_int("AUTO_FILL_MAX_DAYS", 3) if _env_bool("AUTO_FILL_GAPS", True) else 0
    filled = _fill_gaps_before(D, max_days=max_days) if max_days > 0 else []
    payload.date_str = D
    payload.force = True
    payload.push = True
    res = process_and_push(payload, date_str=D, force=True)
    if isinstance(res, dict):
        res["target_date"] = D
        res["gap_filled"] = filled
    return res

@app.post("/api/process_slot")
def process_slot(req: SlotReq):
    target = _prev_trading_day_ymd(datetime.date.today())

    payload = RunReq(date_str=target, force=req.force, push=False, slot=req.slot or "")
    payload.lookback_days = _env_int("MAIL_LOOKBACK_DAYS", 3)
    payload.folder = None
    payload.folders = []
    payload.target_val_date_only = True
    fetch_res = fetch_only(request=None, payload=payload, date_str=target, force=req.force)

    if _env_bool("AUTO_FILL_GAPS", True):
        try:
            _fill_gaps_before(target, max_days=_env_int("AUTO_FILL_MAX_DAYS", 3))
        except Exception:
            pass

    payload2 = RunReq(date_str=target, force=True, push=True, slot=req.slot or "")
    res = process_and_push(payload2, date_str=target, force=True)
    if isinstance(res, dict):
        res["slot"] = req.slot or ""
        res["target_date"] = target
        res["fetch"] = fetch_res
    return res


@app.get("/api/list_folders")
def list_folders():
    # list IMAP folders for debugging
    host=IMAP_HOST
    user=IMAP_USER
    pwd=IMAP_PASS or IMAP_PASSWORD
    from imapclient import IMAPClient
    with IMAPClient(host, ssl=True) as cli:
        cli.login(user, pwd)
        folders = [f[2] for f in cli.list_folders()]
    return {"ok": True, "folders": folders}



def _write_raw_manifest_from_db(date_str: str, out_path: str, base_url: str = ""):
    """
    永远以 DB(raw_ingest) 为准重写清单，避免被空 folder 覆盖成空表
    """
    import os, sqlite3
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    db_path = os.path.join(DATA_DIR, "db.sqlite3")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
      SELECT product_key, date, orig_filename, mail_subject, mail_from, received_at, raw_path
      FROM raw_ingest WHERE date=?
      ORDER BY received_at DESC
    """, (date_str,))
    rows = cur.fetchall()
    conn.close()

    wb=Workbook()
    ws=wb.active
    ws.title="原文件清单"
    headers=["产品名称","日期","原文件名","邮件标题","发件人","收件时间","下载原文件","备注"]
    ws.append(headers)
    bold=Font(bold=True)
    for c in range(1,len(headers)+1):
        ws.cell(row=1,column=c).font=bold

    for (product_key, date, orig_filename, mail_subject, mail_from, received_at, raw_path) in rows:
        dl = ""
        if raw_path:
            # 统一用 download_raw?path=...
            from urllib.parse import quote
            q = quote(raw_path, safe="")
            dl = (base_url.rstrip("/") + "/download_raw?path=" + q) if base_url else ("/download_raw?path=" + q)
        ws.append([product_key, date, orig_filename, mail_subject, mail_from, received_at, "点击下载", ""])
        r=ws.max_row
        if dl:
            ws.cell(r,7).hyperlink = dl
            ws.cell(r,7).font = Font(color="0563C1", underline="single")

    # autosize
    for col in range(1, ws.max_column+1):
        mx=0
        for row in range(1, ws.max_row+1):
            v=ws.cell(row=row,column=col).value
            if v is None: continue
            mx=max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width=min(max(mx+2,10),60)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return {"rows": len(rows), "path": out_path}





def _extract_valuation_date(file_bytes: bytes, fallback_ymd: str) -> str:
    """Try to parse valuation date (YYYYMMDD) from file bytes; fallback to provided ymd."""
    import re
    blob = ""
    for enc in ("utf-8", "latin1"):
        try:
            blob = file_bytes.decode(enc, errors="ignore")
            if blob:
                break
        except Exception:
            pass

    m = re.search(r"(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})", blob)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
        return f"{y}{mo:02d}{d:02d}"

    m = re.search(r"(20\d{2})(\d{2})(\d{2})", blob)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}"

    return fallback_ymd


def _extract_val_date_for_ingest(file_bytes: bytes, fallback_date_str: str, filename: str = "") -> str:
    """Prefer decoded filename date, then parsed valuation date from file content, then fallback."""
    try:
        fn_ymd = _extract_ymd_from_filename(filename or "")
        if fn_ymd:
            return fn_ymd
    except Exception:
        pass
    try:
        vd = _extract_valuation_date(file_bytes, fallback_date_str)
        if vd and len(vd) == 8:
            return vd
    except Exception:
        pass
    return fallback_date_str

def _extract_nav_from_xlsx_bytes(file_bytes: bytes):
    """
    返回 (nav_value: float|None, reason: str)

    适配多模板策略：
    1) 全表扫描定位“单位净值”类关键词（排除累计/复权/万份收益等）
    2) 命中后：
       - 同一单元格内正则抠数（“单位净值 1.2345”）
       - 同行右侧 1~8 格找数
       - 同列下方 1~20 行找数（表头型最常见）
    3) 候选数过滤：0.0001~100；优先“单位净值”>“基金单位净值”>“最新单位净值”
    """
    import io, math, re
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return None, f"openpyxl_load_fail:{e}"

    # 关键词优先级（越靠前越优先）
    kw_priority = [
        "单位净值",
        "基金单位净值",
        "最新单位净值",
        "今日单位净值",
        "份额净值",
        "净值(元)",
        "净值",
    ]
    # 排除词（命中则跳过）
    bad_words = ["累计", "复权", "万份", "收益", "七日年化", "年化", "估值增值", "估值增值率"]

    def to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            x = float(v)
            if math.isfinite(x):
                return x
            return None
        sv = str(v).strip().replace(",", "")
        # 可能含 %，净值不该是百分比；直接排除
        if "%" in sv:
            return None
        # 纯数字
        try:
            return float(sv)
        except Exception:
            return None

    def in_range(x: float) -> bool:
        return x is not None and x > 0.0001 and x < 100

    nav_in_cell_re = re.compile(r"(单位净值|基金单位净值|最新单位净值|份额净值)\s*(?:\(|（)?(?:元)?(?:\)|）)?\s*[:：=]?\s*([0-9]+(?:\.[0-9]+)?)")
    num_re = re.compile(r"([0-9]+(?:\.[0-9]+)?)")

    # 记录所有候选 (priority_idx, score, nav)
    cands = []

    for ws in wb.worksheets[:6]:
        # 扫描范围加大一点，但控制 max_col
        # 注意：values_only=True 可能丢掉一些合并信息，但顶层值仍在；我们靠广扫弥补
        rows = list(ws.iter_rows(min_row=1, max_row=600, max_col=60, values_only=True))
        R = len(rows)

        for i in range(R):
            row = rows[i]
            if not row:
                continue
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                txt = str(cell).strip()
                if not txt:
                    continue

                # 排除干扰项
                if any(bw in txt for bw in bad_words):
                    continue

                # 看是否命中优先级关键词
                pri = None
                for idx, kw in enumerate(kw_priority):
                    if kw in txt:
                        pri = idx
                        break
                if pri is None:
                    continue

                # A) 同格抠数：例如 “单位净值：1.2345”
                m = nav_in_cell_re.search(txt) or num_re.search(txt)
                if m:
                    x = to_float(m.group(2) if getattr(m, "lastindex", 1) and m.lastindex >= 2 else m.group(1))
                    if in_range(x):
                        # score：同格最高
                        cands.append((pri, 0, x, f"ok:in_cell:{kw_priority[pri]}"))
                        continue

                # B) 同行右侧 1~8 格
                for k in range(j + 1, min(j + 9, len(row))):
                    x = to_float(row[k])
                    if in_range(x):
                        cands.append((pri, 1, x, f"ok:right:{kw_priority[pri]}"))
                        break

                # C) 同列下方 1~20 行（表头型）
                for r2 in range(i + 1, min(i + 21, R)):
                    if j >= len(rows[r2]):
                        continue
                    x = to_float(rows[r2][j])
                    if in_range(x):
                        cands.append((pri, 2, x, f"ok:down:{kw_priority[pri]}"))
                        break

    if not cands:
        return None, "nav_not_found"

    # 选最优：优先 pri 小，其次 score 小（同格>右>下），最后数值更“像净值”（1附近优先）
    cands.sort(key=lambda t: (t[0], t[1], abs(t[2]-1.0)))
    pri, score, nav, reason = cands[0]
    return float(nav), reason



def _prev_weekday(ymd: str) -> str:
    import datetime
    dt = datetime.datetime.strptime(ymd, "%Y%m%d").date()
    for _ in range(10):
        dt -= datetime.timedelta(days=1)
        if dt.weekday() < 5:
            return dt.strftime("%Y%m%d")
    return ymd


def _last_friday(ymd: str) -> str:
    import datetime
    dt=datetime.datetime.strptime(ymd,"%Y%m%d").date()
    # 严格取“早于当日”的最近周五（上周五口径）
    dt -= datetime.timedelta(days=1)
    for _ in range(21):
        if dt.weekday() == 4:
            return dt.strftime("%Y%m%d")
        dt -= datetime.timedelta(days=1)
    return ymd



def _prev_month_last_weekday(ymd: str) -> str:
    import datetime
    dt = datetime.datetime.strptime(ymd, "%Y%m%d").date()
    first = dt.replace(day=1)
    d = first - datetime.timedelta(days=1)
    for _ in range(10):
        if d.weekday() < 5:
            return d.strftime("%Y%m%d")
        d -= datetime.timedelta(days=1)
    return ymd


def _year_start_weekday(ymd: str) -> str:
    import datetime
    dt = datetime.datetime.strptime(ymd, "%Y%m%d").date()
    d = dt.replace(month=1, day=1)
    for _ in range(10):
        if d.weekday() < 5:
            return d.strftime("%Y%m%d")
        d += datetime.timedelta(days=1)
    return ymd

def _get_unit_nav(conn, product_key: str, ymd: str):
    cur = conn.cursor()
    where_sql, where_args = _series_where(product_key)
    cur.execute(f"SELECT COALESCE(unit_nav, nav) FROM nav_series WHERE {where_sql} AND date=? AND COALESCE(unit_nav, nav) IS NOT NULL ORDER BY product_key LIMIT 1", (*where_args, ymd))
    r = cur.fetchone()
    return float(r[0]) if r else None


def _calc_day_from_unit_series(conn, product_key: str, ymd: str):
    prev = _prev_weekday(ymd)
    nav_today = _get_unit_nav(conn, product_key, ymd)
    nav_prev = _get_unit_nav(conn, product_key, prev)
    if not nav_today or not nav_prev or nav_prev <= 0:
        return None
    return nav_today / nav_prev - 1.0


def _get_nav(conn, product_key: str, ymd: str):
    cur = conn.cursor()
    where_sql, where_args = _series_where(product_key)
    cur.execute(f"SELECT cum_nav FROM nav_series WHERE {where_sql} AND date=? AND cum_nav IS NOT NULL ORDER BY product_key LIMIT 1", (*where_args, ymd))
    r = cur.fetchone()
    return float(r[0]) if r else None

def _calc_returns_from_series(conn, product_key: str, ymd: str):
    """
    返回 (day, week, month, ytd) 百分比（小数），缺失返回 None
    """
    nav_today=_get_nav(conn, product_key, ymd)
    if not nav_today or nav_today<=0:
        return None,None,None,None

    d_prev=_prev_weekday(ymd)
    d_week=_last_friday(ymd)
    d_month=_prev_month_last_weekday(ymd)

    # YTD anchor: previous year end (12/31), else nearest <= 12/31 with cum_nav
    prev_year_end = f"{int(ymd[:4]) - 1}1231"
    cur = conn.cursor()
    where_sql, where_args = _series_where(product_key)
    cur.execute(
        f"SELECT date FROM nav_series WHERE {where_sql} AND date<=? AND cum_nav IS NOT NULL ORDER BY date DESC LIMIT 1",
        (*where_args, prev_year_end),
    )
    rr = cur.fetchone()
    d_ytd = rr[0] if rr else None

    def calc(base_date):
        if not base_date:
            return None
        nav_base=_get_nav(conn, product_key, base_date)
        if not nav_base or nav_base<=0:
            return None
        return nav_today/nav_base - 1.0

    return calc(d_prev), calc(d_week), calc(d_month), calc(d_ytd)


def _sniff_file_kind(b: bytes) -> str:
    """
    return: xlsx_zip | xls_ole | html | text | unknown
    """
    if not b or len(b) < 8:
        return "unknown"
    # xlsx is zip: PK\x03\x04
    if b[:4] == b"PK\x03\x04":
        return "xlsx_zip"
    # old xls is OLE2: D0 CF 11 E0 A1 B1 1A E1
    if b[:8] == bytes.fromhex("D0CF11E0A1B11AE1"):
        return "xls_ole"
    # html
    head = b[:512].decode("latin1", errors="ignore").lower()
    if "<html" in head or "<table" in head:
        return "html"
    # plain text (csv etc.)
    if all(32 <= x <= 126 or x in (9,10,13) for x in b[:200]):
        return "text"
    return "unknown"


def _extract_nav_any(file_bytes: bytes):
    """
    统一入口：返回 (nav, reason)
    """
    kind = _sniff_file_kind(file_bytes)

    # a) xlsx zip -> openpyxl 走你现有的 _extract_nav_from_xlsx_bytes
    if kind == "xlsx_zip":
        nav, reason = _extract_nav_from_xlsx_bytes(file_bytes)
        return nav, reason

    # b) xls ole -> xlrd
    if kind == "xls_ole":
        try:
            import xlrd, io
            book = xlrd.open_workbook(file_contents=file_bytes)
            kws = NAV_FIELD_KEYWORDS or ["单位净值","基金单位净值","最新单位净值","份额净值"]
            for si in range(min(3, book.nsheets)):
                sh = book.sheet_by_index(si)
                R = min(sh.nrows, 600)
                C = min(sh.ncols, 60)
                for r in range(R):
                    for c in range(C):
                        v = sh.cell_value(r,c)
                        if isinstance(v,str) and any(k in v for k in kws):
                            # 同格抠数
                            import re
                            m = re.search(r"([0-9]+(?:\.[0-9]+)?)", v.replace("：",":"))
                            if m:
                                return float(m.group(1)), "ok:xls_in_cell"
                            # 右侧
                            for k in range(c+1, min(c+9, C)):
                                vv = sh.cell_value(r,k)
                                if isinstance(vv,(int,float)) and 0.0001 < float(vv) < 100:
                                    return float(vv), "ok:xls_right"
                                if isinstance(vv,str):
                                    mm = re.search(r"([0-9]+(?:\.[0-9]+)?)", vv)
                                    if mm:
                                        x=float(mm.group(1))
                                        if 0.0001 < x < 100:
                                            return x, "ok:xls_right_str"
                            # 下方
                            for k in range(r+1, min(r+21, R)):
                                vv = sh.cell_value(k,c)
                                if isinstance(vv,(int,float)) and 0.0001 < float(vv) < 100:
                                    return float(vv), "ok:xls_down"
                                if isinstance(vv,str):
                                    mm = re.search(r"([0-9]+(?:\.[0-9]+)?)", vv)
                                    if mm:
                                        x=float(mm.group(1))
                                        if 0.0001 < x < 100:
                                            return x, "ok:xls_down_str"
            return None, "nav_not_found_xls"
        except Exception as e:
            return None, f"xls_parse_fail:{e}"

    # c) html -> pandas read_html
    if kind == "html":
        try:
            import pandas as pd
            from io import BytesIO
            html = file_bytes.decode("utf-8", errors="ignore")
            tables = pd.read_html(html)
            kws = NAV_FIELD_KEYWORDS or ["单位净值","基金单位净值","最新单位净值","份额净值"]
            import re
            for df in tables[:5]:
                # 扫描每个 cell
                for r in range(min(len(df), 400)):
                    for c in range(min(len(df.columns), 60)):
                        v = df.iat[r,c]
                        if v is None: 
                            continue
                        sv = str(v)
                        if any(k in sv for k in kws):
                            # 同格抠数
                            m = re.search(r"([0-9]+(?:\.[0-9]+)?)", sv.replace("：",":"))
                            if m:
                                x=float(m.group(1))
                                if 0.0001 < x < 100:
                                    return x, "ok:html_in_cell"
                            # 右侧/下方
                            if c+1 < len(df.columns):
                                vv=str(df.iat[r,c+1])
                                m2=re.search(r"([0-9]+(?:\.[0-9]+)?)", vv)
                                if m2:
                                    x=float(m2.group(1))
                                    if 0.0001 < x < 100:
                                        return x, "ok:html_right"
                            if r+1 < len(df):
                                vv=str(df.iat[r+1,c])
                                m2=re.search(r"([0-9]+(?:\.[0-9]+)?)", vv)
                                if m2:
                                    x=float(m2.group(1))
                                    if 0.0001 < x < 100:
                                        return x, "ok:html_down"
            return None, "nav_not_found_html"
        except Exception as e:
            return None, f"html_parse_fail:{e}"

    return None, f"unsupported_kind:{kind}"


def _norm_ymd_any(x):
    import re, datetime
    if x is None: return None
    s=str(x).strip()
    if not s: return None
    # 2026-03-06 / 2026/03/06
    m=re.search(r"(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})", s)
    if m:
        y=int(m.group(1)); mo=int(m.group(2)); d=int(m.group(3))
        return f"{y}{mo:02d}{d:02d}"
    # 20260306
    m=re.search(r"(20\d{2})(\d{2})(\d{2})", s)
    if m:
        return m.group(1)+m.group(2)+m.group(3)
    return None

def _to_float_nav(v):
    import math, re
    if v is None: return None
    if isinstance(v,(int,float)):
        x=float(v)
        if math.isfinite(x) and 0.0001 < x < 100: return x
        return None
    s=str(v).strip().replace(",","")
    if not s or "%" in s: return None
    m=re.search(r"([0-9]+(?:\.[0-9]+)?)", s)
    if not m: return None
    try:
        x=float(m.group(1))
        if 0.0001 < x < 100: return x
    except:
        return None
    return None


def upsert_nav2(conn, product_key: str, date_str: str, unit_nav: float | None, cum_nav: float | None):
    """
    Write nav_series with both unit_nav and cum_nav.
    Keep legacy nav column for compatibility: nav := cum_nav if exists else unit_nav.
    """
    nav_val = cum_nav if cum_nav is not None else unit_nav
    conn.execute(
        """
        INSERT INTO nav_series(product_key,date,nav,unit_nav,cum_nav)
        VALUES(?,?,?,?,?)
        ON CONFLICT(product_key,date) DO UPDATE SET
            nav=excluded.nav,
            unit_nav=COALESCE(excluded.unit_nav, nav_series.unit_nav),
            cum_nav=COALESCE(excluded.cum_nav, nav_series.cum_nav)
        """,
        (product_key, date_str, nav_val, unit_nav, cum_nav),
    )


def _import_nav_history_from_bytes(file_bytes: bytes, fallback_product_key: str = "", filename: str = ""):
    """
    Import history series into nav_series with unit_nav + cum_nav.
    Supported formats: xlsx/xls/html.
    Product key priority:
      1) fallback_product_key
      2) filename fuzzy match against existing nav_series/raw_ingest keys
    """
    import io
    import os
    import re
    import pandas as pd

    def _norm_text(x) -> str:
        s = str(x or '').strip()
        s = s.replace('　', ' ').replace('_', '')
        s = s.replace('（', '(').replace('）', ')').replace('：', ':')
        s = re.sub(r'\s+', '', s)
        return s

    def _resolve_history_product_key(name: str, hint: str = '') -> str:
        hint = (hint or '').strip()
        if hint:
            return hint
        base = os.path.splitext(os.path.basename(name or ''))[0]
        base_norm = _norm_text(base)
        # trim common prefixes/suffixes from manual history filenames
        base_norm = re.sub(r'^净值序列\d{8}\d{6}', '', base_norm)
        base_norm = re.sub(r'^净值序列', '', base_norm)
        base_norm = re.sub(r'净值波动表20\d{6}至20\d{6}\d*$', '', base_norm)
        base_norm = re.sub(r'净值波动表20\d{6}至20\d{6}_?\d*$', '', base_norm)
        base_norm = re.sub(r'_?\d{12,}$', '', base_norm)
        code_m = re.match(r'^([A-Z]{2,5}\d{2,6})', base_norm)
        code = code_m.group(1) if code_m else ''

        with connect(DATA_DIR) as conn:
            cur = conn.cursor()
            keys = set()
            for row in cur.execute('SELECT DISTINCT product_key FROM nav_series'):
                if row[0]:
                    keys.add(row[0])
            for row in cur.execute('SELECT DISTINCT product_key FROM raw_ingest'):
                if row[0]:
                    keys.add(row[0])
        candidates = []
        for pk in keys:
            pk_norm = _norm_text(pk)
            score = 0
            if code and pk_norm.startswith(code):
                score += 100
            if base_norm and base_norm in pk_norm:
                score += 80 + len(base_norm)
            if pk_norm and pk_norm in base_norm:
                score += 40 + len(pk_norm)
            # strip noisy suffix for Chinese-name overlap
            slim = re.sub(r'委托资产资产|私募证券投资基金|私募证券投资|证券投资基金|证券投资', '', base_norm)
            if slim and slim in pk_norm:
                score += 30 + len(slim)
            if score > 0:
                candidates.append((score, len(pk_norm), pk))
        candidates.sort(reverse=True)
        return candidates[0][2] if candidates else base

    def _col_kind(name: str) -> str | None:
        n = _norm_text(name)
        if n in {'净值日期', '估值日期', '日期'}:
            return 'date'
        if n in {'单位净值', '单位净值(元)', '单位净值元', '基金单位净值', '份额净值', '最新单位净值', '今日单位净值'}:
            return 'unit'
        if n in {'累计净值', '累计净值(元)', '累计净值元', '累计单位净值', '累计单位净值(元)', '复权净值'}:
            return 'cum'
        if '单位净值' in n and '累计' not in n:
            return 'unit'
        if '累计' in n and '净值' in n:
            return 'cum'
        if '复权净值' in n:
            return 'cum'
        return None

    frames = []
    fn = (filename or '').lower()
    try:
        if fn.endswith(('.html', '.htm')):
            frames = pd.read_html(file_bytes)
        else:
            bio = io.BytesIO(file_bytes)
            book = pd.read_excel(bio, sheet_name=None, engine=None, header=None)
            frames = list(book.values())
    except Exception:
        return 0

    product_key = _resolve_history_product_key(filename, fallback_product_key)
    upserts = 0
    with connect(DATA_DIR) as conn:
        for df in frames:
            if df is None or df.empty:
                continue
            header_row = None
            date_col = None
            unit_col = None
            cum_col = None
            max_r = min(len(df), 30)
            max_c = min(len(df.columns), 40)
            for r in range(max_r):
                found = {}
                for c in range(max_c):
                    kind = _col_kind(df.iat[r, c])
                    if kind and kind not in found:
                        found[kind] = c
                if 'date' in found and ('unit' in found or 'cum' in found):
                    header_row = r
                    date_col = found.get('date')
                    unit_col = found.get('unit')
                    cum_col = found.get('cum')
                    break
            if header_row is None:
                continue
            for r in range(header_row + 1, min(len(df), header_row + 2500)):
                dv = df.iat[r, date_col] if date_col is not None else None
                ymd = _norm_ymd_any(dv)
                if not ymd:
                    continue
                unit_v = _to_float_nav(df.iat[r, unit_col]) if unit_col is not None else None
                cum_v = _to_float_nav(df.iat[r, cum_col]) if cum_col is not None else None
                if unit_v is None and cum_v is None:
                    continue
                upsert_nav2(conn, product_key, ymd, unit_v, cum_v)
                upserts += 1
        conn.commit()
    print(f"[IMPORT_HISTORY_OK] product_key={product_key} filename={filename or ''} upserts={upserts}", flush=True)
    return upserts

@app.post("/api/upload_history")
async def upload_history(files: list[UploadFile] = File(...)):
    """
    上传历史净值表（可多文件），识别并导入 nav_series。
    返回每个文件导入条数 + 总条数。
    """
    total = 0
    items = []
    for f in files:
        name = f.filename or "unknown"
        b = await f.read()
        try:
            n = _import_nav_history_from_bytes(b, fallback_product_key="", filename=name)
            total += int(n or 0)
            items.append({"filename": name, "bytes": len(b), "upserts": int(n or 0), "ok": True})
        except Exception as e:
            items.append({"filename": name, "bytes": len(b), "upserts": 0, "ok": False, "error": str(e)})
    return {"ok": True, "total_upserts": total, "items": items}



@app.post("/api/import_dir_history")
def import_dir_history(payload: dict):
    """
    Import history files already on server directory.
    payload:
      - dir: str
      - product_key_hint: optional
    """
    import os
    import glob
    import zipfile

    dirp = (payload.get("dir") or "").strip()
    if not dirp or not os.path.isdir(dirp):
        return {"ok": False, "error": f"dir not found: {dirp}"}
    pk_hint = (payload.get("product_key_hint") or "").strip() or None

    files = []
    for ext in ("*.xlsx", "*.xls", "*.html", "*.htm", "*.zip"):
        files += glob.glob(os.path.join(dirp, ext))
    files = sorted(set(files))

    total = 0
    per = []
    for fp in files:
        name = os.path.basename(fp)
        if name.lower().endswith('.zip'):
            ztmp = os.path.join(dirp, f'_unz_{name}')
            os.makedirs(ztmp, exist_ok=True)
            with zipfile.ZipFile(fp) as z:
                z.extractall(ztmp)
            inner = []
            for ext in ("*.xlsx", "*.xls", "*.html", "*.htm"):
                inner += glob.glob(os.path.join(ztmp, '**', ext), recursive=True)
            inner = sorted(set(inner))
            sub = 0
            for ip in inner:
                b = open(ip, 'rb').read()
                sub += _import_nav_history_from_bytes(b, fallback_product_key=pk_hint or '', filename=os.path.basename(ip))
            per.append({"file": name, "type": "zip", "inner": len(inner), "upserts": sub})
            total += sub
        else:
            b = open(fp, 'rb').read()
            up = _import_nav_history_from_bytes(b, fallback_product_key=pk_hint or '', filename=name)
            per.append({"file": name, "upserts": up})
            total += up
    return {"ok": True, "dir": dirp, "files": per, "total_upserts": total}

@app.get("/api/nav_series_stats")
def nav_series_stats(limit: int = 50):
    """
    查看 nav_series 统计：每个产品条数、最早/最晚日期
    """
    import os, sqlite3
    db_path = os.path.join(DATA_DIR, "db.sqlite3")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        """
      SELECT product_key, COUNT(*), MIN(date), MAX(date)
      FROM nav_series GROUP BY product_key
      ORDER BY MAX(date) DESC LIMIT ?
    """,
        (limit,),
    )
    rows = cur.fetchall()
    conn.close()
    items = [{"product_key": r[0], "count": r[1], "min_date": r[2], "max_date": r[3]} for r in rows]
    return {"ok": True, "items": items}


@app.get("/api/products")
def api_products():
    with connect(DATA_DIR) as conn_obj:
        return {"ok": True, "items": _get_allowlist_from_db(conn_obj)}


@app.post("/api/products/add")
def api_products_add(req: ProductReq):
    code = _norm_code(req.code)
    if not code:
        raise HTTPException(400, "empty code")
    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO product_allowlist(code, display_name, enabled, updated_at) VALUES(?,?,1,datetime('now'))",
            (code, (req.display_name or "").strip() or dict(_default_target_products()).get(code, code)),
        )
        conn_obj.commit()
    return {"ok": True, "code": code}


@app.post("/api/products/delete")
def api_products_delete(req: ProductReq):
    code = _norm_code(req.code)
    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()
        cur.execute("DELETE FROM product_allowlist WHERE code=?", (code,))
        conn_obj.commit()
    return {"ok": True, "code": code}


@app.post("/api/products/toggle")
def api_products_toggle(req: ToggleReq):
    code = _norm_code(req.code)
    with connect(DATA_DIR) as conn_obj:
        cur = conn_obj.cursor()
        cur.execute(
            "UPDATE product_allowlist SET enabled=?, updated_at=datetime('now') WHERE code=?",
            (1 if req.enabled else 0, code),
        )
        conn_obj.commit()
    return {"ok": True, "code": code, "enabled": req.enabled}


@app.get("/api/raw_ingest_list")
def raw_ingest_list(date_str: str, limit: int = 200):
    """List raw_ingest rows for one day for explainability checks."""
    import os
    import sqlite3
    db_path = os.path.join(DATA_DIR, "db.sqlite3")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        """
      SELECT product_key, date, orig_filename, mail_subject, mail_from, received_at, raw_path
      FROM raw_ingest
      WHERE date = ?
      ORDER BY received_at DESC
      LIMIT ?
    """,
        (date_str, limit),
    )
    rows = cur.fetchall()
    conn.close()
    items = []
    for r in rows:
        items.append({
            "product_key": r[0],
            "date": r[1],
            "orig_filename": r[2],
            "mail_subject": r[3],
            "mail_from": r[4],
            "received_at": r[5],
            "raw_path": r[6],
        })
    return {"ok": True, "date": date_str, "count": len(items), "items": items}


@app.get("/api/nav_series_dump")
def nav_series_dump(product_key: str, limit: int = 50):
    import os
    import sqlite3
    db_path = os.path.join(DATA_DIR, "db.sqlite3")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(
        """
      SELECT date, nav FROM nav_series
      WHERE product_key=?
      ORDER BY date DESC
      LIMIT ?
    """,
        (product_key, limit),
    )
    rows = cur.fetchall()
    conn.close()
    return {
        "ok": True,
        "product_key": product_key,
        "count": len(rows),
        "items": [{"date": r[0], "nav": r[1]} for r in rows],
    }
