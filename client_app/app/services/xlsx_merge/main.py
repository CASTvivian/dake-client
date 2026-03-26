from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sse_starlette.sse import EventSourceResponse

import pandas as pd
import openpyxl
import io, re, os, datetime, uuid, asyncio, time
from urllib.parse import quote
from typing import List, Dict, Any, Optional

app = FastAPI(title="xlsx-merge", version="2.3.1")

# === CLIENT_CONFIG_BRIDGE_BEGIN ===
import json
SERVICE_CONFIG = os.getenv("SERVICE_CONFIG", "").strip()
def _load_client_cfg():
    if not SERVICE_CONFIG or not os.path.exists(SERVICE_CONFIG):
        return None
    try:
        with open(SERVICE_CONFIG, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None
def _apply_client_cfg(cfg: dict):
    data_root = (cfg or {}).get("data_root", "") if cfg else ""
    if data_root:
        os.environ["DATA_DIR"] = os.path.join(data_root, "6000", "xlsx_merge")
@app.post("/api/reload_config")
def reload_config():
    cfg = _load_client_cfg()
    if cfg:
        _apply_client_cfg(cfg)
    return {"ok": True, "loaded": bool(cfg)}
# === CLIENT_CONFIG_BRIDGE_END ===

app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")

HOLD_START_KEYS = ["证券代码","证券名称"]

# =========================
# 解析/合并：你现有逻辑（基本不改）
# =========================
def to_float(x):
    if x is None: return None
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip()
    if s == "" or s.lower() == "nan": return None
    s = s.replace(",", "")
    if s.endswith("%"):
        try: return float(s[:-1]) / 100.0
        except: return None
    try: return float(s)
    except: return None

def load_sheet_raw_bytes(content: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sh = wb[wb.sheetnames[0]]
    data = []
    for r in range(1, sh.max_row + 1):
        data.append([sh.cell(r, c).value for c in range(1, sh.max_column + 1)])
    return pd.DataFrame(data)

def find_row_with(df: pd.DataFrame, keyword: str) -> Optional[int]:
    for i in range(df.shape[0]):
        if keyword in df.iloc[i].tolist():
            return i
    return None

def normalize_code(code) -> Optional[str]:
    if code is None: return None
    s = str(code).strip()
    if s == "" or s.lower() == "nan": return None
    return s  # 保留前导0

def parse_one(content: bytes, account_name: str) -> Dict[str, Any]:
    df = load_sheet_raw_bytes(content)
    header_row = find_row_with(df, "币种")
    if header_row is None:
        raise ValueError(f"[{account_name}] 找不到 '币种' 表头")

    header_vals = df.iloc[header_row].tolist()

    hold_start = None
    for j, v in enumerate(header_vals):
        if v in HOLD_START_KEYS:
            hold_start = j
            break

    # 结构B：同一行表头含现金+持仓
    if hold_start is not None and hold_start > 0 and header_vals[0] == "币种":
        cash_header = header_vals[:hold_start]
        hold_header = header_vals[hold_start:]
        cash_range = list(range(0, hold_start))
        hold_range = list(range(hold_start, len(header_vals)))
        data_start = header_row + 1

        cash_meta = {}
        first_cash = None
        for i in range(data_start, df.shape[0]):
            row = df.iloc[i, cash_range].tolist()
            if not any(x is not None and str(x).strip() != "" for x in row):
                continue
            tmp = {str(k): v for k, v in zip(cash_header, row) if k is not None}
            if first_cash is None:
                first_cash = tmp
            if str(tmp.get("币种", "")).strip() == "人民币":
                cash_meta = tmp
                break
        if not cash_meta:
            cash_meta = first_cash or {}

        holdings_raw = []
        for i in range(data_start, df.shape[0]):
            row = df.iloc[i, hold_range].tolist()
            if not any(x is not None and str(x).strip() != "" for x in row):
                continue
            rec = {"account": account_name}
            for k, val in zip(hold_header, row):
                if k is None: continue
                rec[str(k)] = val
            holdings_raw.append(rec)

        return {"account": account_name, "cash_meta": cash_meta, "holdings_raw": holdings_raw}

    # 结构A：现金在前，持仓在后
    cash_header = header_vals
    cash_range = list(range(0, len(cash_header)))

    cash_meta = {}
    first_cash = None
    for i in range(header_row + 1, df.shape[0]):
        row = df.iloc[i, cash_range].tolist()
        if not any(x is not None and str(x).strip() != "" for x in row):
            continue
        tmp = {str(k): v for k, v in zip(cash_header, row) if k is not None}
        if first_cash is None:
            first_cash = tmp
        if str(tmp.get("币种", "")).strip() == "人民币":
            cash_meta = tmp
            break
    if not cash_meta:
        cash_meta = first_cash or {}

    hold_header_row = None
    for i in range(header_row + 1, df.shape[0]):
        vals = df.iloc[i].tolist()
        if "证券名称" in vals and ("证券数量" in vals or "证券代码" in vals):
            hold_header_row = i
            break

    holdings_raw = []
    if hold_header_row is not None:
        hold_header = df.iloc[hold_header_row].tolist()
        hold_range = list(range(0, len(hold_header)))
        for i in range(hold_header_row + 1, df.shape[0]):
            row = df.iloc[i, hold_range].tolist()
            if not any(x is not None and str(x).strip() != "" for x in row):
                continue
            rec = {"account": account_name}
            for k, val in zip(hold_header, row):
                if k is None: continue
                rec[str(k)] = val
            holdings_raw.append(rec)

    return {"account": account_name, "cash_meta": cash_meta, "holdings_raw": holdings_raw}

def extract_available_cash(cash_meta: Dict[str, Any]) -> float:
    return float(to_float(cash_meta.get("可用")) or 0.0)

def normalize_holding(rec: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "code": normalize_code(rec.get("证券代码") or rec.get("代码") or rec.get("股票代码")),
        "name": (rec.get("证券名称") or rec.get("名称")),
        "qty": float(to_float(rec.get("证券数量")) or 0.0),
        "sellable": float(to_float(rec.get("可卖数量")) or 0.0),
        "mv": float(to_float(rec.get("最新市值")) or 0.0),
    }

def merge_holdings(parsed_list: List[Dict[str, Any]]) -> pd.DataFrame:
    agg: Dict[str, Dict[str, Any]] = {}
    for item in parsed_list:
        for raw in item["holdings_raw"]:
            h = normalize_holding(raw)
            code = h["code"]
            if not code:
                continue
            if code not in agg:
                agg[code] = {
                    "证券代码": code,
                    "证券名称": h["name"],
                    "证券数量": 0.0,
                    "可卖数量": 0.0,
                    "最新市值": 0.0,
                }
            agg[code]["证券数量"] += h["qty"]
            agg[code]["可卖数量"] += h["sellable"]
            agg[code]["最新市值"] += h["mv"]
    df = pd.DataFrame(list(agg.values()))
    if df.empty:
        df = pd.DataFrame(columns=["证券代码","证券名称","证券数量","可卖数量","最新市值"])
    df = df.sort_values(by="最新市值", ascending=False, kind="mergesort").reset_index(drop=True)
    return df

def build_single_sheet(parsed_list: List[Dict[str, Any]], target_amount: float) -> Dict[str, Any]:
    cash_items = []
    cash_sum = 0.0
    for item in parsed_list:
        acct = item["account"]
        avail = extract_available_cash(item["cash_meta"])
        cash_items.append({"name": acct, "available": avail})
        cash_sum += avail

    df_hold = merge_holdings(parsed_list)
    mv_sum = float(df_hold["最新市值"].sum()) if len(df_hold) else 0.0
    total_asset = mv_sum + cash_sum

    if total_asset > 0 and len(df_hold):
        df_hold["持仓比例"] = df_hold["最新市值"] / total_asset
        df_hold["按500w比例金额"] = df_hold["持仓比例"] * float(target_amount)
    else:
        df_hold["持仓比例"] = None
        df_hold["按500w比例金额"] = None

    df_out = df_hold[["证券代码","证券名称","证券数量","可卖数量","最新市值","持仓比例","按500w比例金额"]].copy()

    # 现金占比 & 现金对应金额
    holdings_ratio_sum = float(df_out["持仓比例"].sum()) if len(df_out) and total_asset > 0 else 0.0
    cash_ratio = max(0.0, 1.0 - holdings_ratio_sum) if total_asset > 0 else 0.0
    holdings_amount_sum = float(df_out["按500w比例金额"].sum()) if len(df_out) else 0.0
    cash_amount = max(0.0, float(target_amount) - holdings_amount_sum)

    return {
        "df_out": df_out,
        "cash_items": cash_items,
        "mv_sum": mv_sum,
        "cash_sum": cash_sum,
        "total_asset": total_asset,
        "cash_ratio": cash_ratio,
        "cash_amount": cash_amount,
        "holdings_ratio_sum": holdings_ratio_sum,
        "holdings_amount_sum": holdings_amount_sum,
    }

def safe_sheet_name(name: str) -> str:
    s = re.sub(r"[\[\]\:\*\?\/\\]+", "_", name)
    return (s[:31] or "Sheet1")

def resolve_date(date_str: Optional[str]) -> str:
    if date_str:
        d = re.sub(r"[^0-9]", "", date_str.strip())
        if len(d) == 8:
            return d
    return datetime.datetime.now().strftime("%Y%m%d")

def make_excel_bytes(built: Dict[str, Any], date_yyyymmdd: str, target_amount: float) -> bytes:
    df_out = built["df_out"]
    cash_items = built["cash_items"]

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        sh = safe_sheet_name(date_yyyymmdd)
        df_out.to_excel(writer, sheet_name=sh, index=False, startrow=0, startcol=0)

        workbook = writer.book
        worksheet = writer.sheets[sh]

        fmt_money = workbook.add_format({'num_format': '#,##0.00'})
        fmt_pct = workbook.add_format({'num_format': '0.0000%'})
        fmt_title = workbook.add_format({'bold': True})

        worksheet.set_column('A:A', 12)
        worksheet.set_column('B:B', 16)
        worksheet.set_column('C:D', 12, fmt_money)
        worksheet.set_column('E:E', 14, fmt_money)
        worksheet.set_column('F:F', 12, fmt_pct)
        worksheet.set_column('G:G', 16, fmt_money)
        worksheet.freeze_panes(1, 0)

        # 表尾追加区块
        start_row = len(df_out) + 3
        worksheet.write(start_row, 3, "账户可用现金", fmt_title)
        r = start_row + 1
        for item in cash_items:
            worksheet.write(r, 3, item["name"])
            worksheet.write_number(r, 4, float(item["available"] or 0.0), fmt_money)
            r += 1

        r += 1
        worksheet.write(r, 3, "A股总资产", fmt_title)
        worksheet.write_number(r, 4, float(built["total_asset"] or 0.0), fmt_money)
        worksheet.write_number(r, 6, float(target_amount), fmt_money)

        r += 2
        worksheet.write(r, 3, "现金占比", fmt_title)
        worksheet.write_number(r, 4, float(built["cash_ratio"] or 0.0), fmt_pct)
        worksheet.write(r+1, 3, "现金对应金额", fmt_title)
        worksheet.write_number(r+1, 4, float(built["cash_amount"] or 0.0), fmt_money)

    out.seek(0)
    return out.read()

# =========================
# 首页/健康检查
# =========================
@app.get("/")
def home():
    return FileResponse(os.path.join(os.path.dirname(__file__), "static", "index.html"))

@app.get("/health")
def health():
    return {"ok": True, "service": "xlsx-merge", "version": "2.3.1", "ts": datetime.datetime.now().isoformat()}

# =========================
# 旧接口：仍保留（直接下载）
# =========================
@app.post("/api/merge")
async def merge_direct(
    files: List[UploadFile] = File(..., alias="files[]"),
    target_amount: float = Form(5000000),
    date_str: Optional[str] = Form(None),
):
    parsed = []
    for f in files:
        content = await f.read()
        acct = os.path.splitext(os.path.basename(f.filename))[0]
        acct = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fa5_]+", "_", acct)
        parsed.append(parse_one(content, acct))

    d = resolve_date(date_str)

    async with _jobs_lock:
        job = _jobs.get(job_id)
        if job:
            job["date"] = d
    built = build_single_sheet(parsed, target_amount=target_amount)
    file_bytes = make_excel_bytes(built, d, target_amount)

    filename_cn = f"持仓比例-{d}.xlsx"
    filename_ascii = f"holding_ratio-{d}.xlsx"
    content_disposition = f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{quote(filename_cn)}'

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition},
    )

@app.post("/api/preview")
async def preview(
    files: List[UploadFile] = File(..., alias="files[]"),
    target_amount: float = Form(5000000),
):
    parsed = []
    for f in files:
        content = await f.read()
        acct = os.path.splitext(os.path.basename(f.filename))[0]
        acct = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fa5_]+", "_", acct)
        parsed.append(parse_one(content, acct))

    built = build_single_sheet(parsed, target_amount=target_amount)
    return JSONResponse({
        "cash_items": built["cash_items"],
        "mv_sum": built["mv_sum"],
        "cash_sum": built["cash_sum"],
        "total_asset": built["total_asset"],
        "holdings_ratio_sum": built["holdings_ratio_sum"],
        "cash_ratio": built["cash_ratio"],
        "holdings_amount_sum": built["holdings_amount_sum"],
        "cash_amount": built["cash_amount"],
        "holdings": built["df_out"].to_dict(orient="records"),
    })

# =========================
# 新标准：Job + SSE
# =========================
JOB_TTL_SECONDS = 10 * 60
MAX_JOBS = 200  # 防刷：最多保留200个job（TTL会自动清理）  # 10分钟自动清理（内存无留痕）
_jobs: Dict[str, Dict[str, Any]] = {}
_jobs_lock = asyncio.Lock()

async def _job_emit(job_id: str, payload: Dict[str, Any]):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        job["last_emit"] = time.time()
        q: asyncio.Queue = job["queue"]
    await q.put(payload)

async def _job_cleanup_later(job_id: str):
    await asyncio.sleep(JOB_TTL_SECONDS)
    async with _jobs_lock:
        _jobs.pop(job_id, None)

def _mk_cd(d: str):
    filename_cn = f"持仓比例-{d}.xlsx"
    filename_ascii = f"holding_ratio-{d}.xlsx"
    cd = f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{quote(filename_cn)}'
    return filename_cn, filename_ascii, cd

async def _run_job(job_id: str, files_payload: List[Dict[str, Any]], target_amount: float, date_str: Optional[str]):
    """
    files_payload: [{name, content_bytes}]
    在内存完成解析 -> 合并 -> 生成 xlsx bytes
    """
    d = resolve_date(date_str)

    async with _jobs_lock:
        job = _jobs.get(job_id)
        if job:
            job["date"] = d

    try:
        await _job_emit(job_id, {"stage": "start", "progress": 1, "message": "开始处理"})

        parsed: List[Dict[str, Any]] = []
        total_files = len(files_payload) or 1

        # 解析阶段：给真实进度（10%~70%）
        for idx, fp in enumerate(files_payload, start=1):
            # cancel check
            async with _jobs_lock:
                job = _jobs.get(job_id)
                if not job:
                    return
                if job.get("cancelled"):
                    raise RuntimeError("已取消")

            acct = fp["name"]
            content = fp["content"]
            await _job_emit(job_id, {
                "stage": "parse",
                "progress": int(10 + (idx-1)/total_files * 60),
                "message": f"解析文件 {idx}/{total_files}: {acct}"
            })

            # openpyxl/pandas 重，放到线程里
            item = await asyncio.to_thread(parse_one, content, acct)
            parsed.append(item)

        await _job_emit(job_id, {"stage": "merge", "progress": 75, "message": "聚合持仓 & 计算比例"})
        built = await asyncio.to_thread(build_single_sheet, parsed, target_amount)

        await _job_emit(job_id, {"stage": "write", "progress": 88, "message": "生成 Excel"})
        xlsx_bytes = await asyncio.to_thread(make_excel_bytes, built, d, target_amount)

        filename_cn, filename_ascii, cd = _mk_cd(d)

        async with _jobs_lock:
            job = _jobs.get(job_id)
            if not job:
                return
            if job.get("cancelled"):
                raise RuntimeError("已取消")
            job["status"] = "done"
            job["result_bytes"] = xlsx_bytes
            job["filename_cn"] = filename_cn
            job["filename_ascii"] = filename_ascii
            job["content_disposition"] = cd

        await _job_emit(job_id, {"stage": "done", "progress": 100, "message": "完成，可下载"})
        asyncio.create_task(_job_cleanup_later(job_id))

    except Exception as e:
        async with _jobs_lock:
            job = _jobs.get(job_id)
            if job:
                job["status"] = "error"
                job["error"] = str(e)
        await _job_emit(job_id, {"stage": "error", "progress": 0, "message": str(e)})
        asyncio.create_task(_job_cleanup_later(job_id))

@app.post("/api/merge_job")
async def merge_job(
    files: List[UploadFile] = File(..., alias="files[]"),
    target_amount: float = Form(5000000),
    date_str: Optional[str] = Form(None),
):
    # 防刷：限制 job 数量（TTL 自动清理）
    async with _jobs_lock:
        if len(_jobs) >= MAX_JOBS:
            raise HTTPException(status_code=429, detail='too many jobs, try later')

    # 读取上传内容到内存（无落盘）
    files_payload = []
    for f in files:
        content = await f.read()
        name = os.path.splitext(os.path.basename(f.filename))[0]
        name = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fa5_]+", "_", name)
        files_payload.append({"name": name, "content": content})

    job_id = uuid.uuid4().hex
    q: asyncio.Queue = asyncio.Queue()

    async with _jobs_lock:
        _jobs[job_id] = {
            "status": "running",
            "queue": q,
            "created": time.time(),
            "last_emit": time.time(),
            "cancelled": False,
            "result_bytes": None,
            "error": None,
        }

    # 启动后台任务
    asyncio.create_task(_run_job(job_id, files_payload, target_amount, date_str))

    return JSONResponse({
        "job_id": job_id,
        "progress_url": f"/api/progress/{job_id}",
        "download_url": f"/api/download/{job_id}",
        "download_blob_url": f"/api/download_blob/{job_id}",
        "job_url": f"/api/job/{job_id}",
        "cancel_url": f"/api/cancel/{job_id}",
    })

@app.get("/api/progress/{job_id}")
async def progress(job_id: str):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job not found")
        q: asyncio.Queue = job["queue"]

    async def event_gen():
        # 先发一条“连接成功”
        yield {"event": "message", "data": {"stage": "connected", "progress": 0, "message": "已连接"}}
        while True:
            try:
                payload = await asyncio.wait_for(q.get(), timeout=10)
                yield {"event": "message", "data": payload}

                # done/error 就结束 SSE
                if payload.get("stage") in ("done", "error"):
                    break
            except asyncio.TimeoutError:
                # keep-alive
                yield {"event": "ping", "data": {"ts": int(time.time())}}

    return EventSourceResponse(event_gen())

@app.get("/api/download/{job_id}")
async def download(job_id: str):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job not found")
        if job.get("status") == "error":
            raise HTTPException(status_code=400, detail=job.get("error") or "error")
        if job.get("status") != "done" or not job.get("result_bytes"):
            raise HTTPException(status_code=409, detail="not ready")
        b = job["result_bytes"]
        cd = job["content_disposition"]

    return StreamingResponse(
        io.BytesIO(b),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )

@app.get("/api/download_blob/{job_id}")
async def download_blob(job_id: str):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job not found")
        if job.get("status") == "error":
            raise HTTPException(status_code=400, detail=job.get("error") or "error")
        if job.get("status") != "done" or not job.get("result_bytes"):
            raise HTTPException(status_code=409, detail="not ready")
        b = job["result_bytes"]
    return StreamingResponse(
        io.BytesIO(b),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@app.get("/api/job/{job_id}")
async def job_status(job_id: str):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job not found")
        return {
            "job_id": job_id,
            "status": job.get("status"),
            "error": job.get("error"),
            "date": job.get("date"),
            "has_result": bool(job.get("result_bytes")),
            "download_url": f"/api/download/{job_id}",
            "download_blob_url": f"/api/download_blob/{job_id}",
            "job_url": f"/api/job/{job_id}",
            "progress_url": f"/api/progress/{job_id}",
            "cancel_url": f"/api/cancel/{job_id}",
        }

@app.post("/api/cancel/{job_id}")
async def cancel(job_id: str):
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="job not found")
        job["cancelled"] = True
        job["status"] = "error"
        job["error"] = "已取消"
        q: asyncio.Queue = job["queue"]

    # 通知 SSE 结束
    await q.put({"stage": "error", "progress": 0, "message": "已取消"})
    asyncio.create_task(_job_cleanup_later(job_id))
    return JSONResponse({"ok": True, "job_id": job_id, "status": "cancelled"})
